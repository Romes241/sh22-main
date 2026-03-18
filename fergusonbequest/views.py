from datetime import datetime, timedelta
from django import forms
from django.conf import settings
from django.contrib.auth import get_user_model, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import LoginView
from django.core.mail import send_mail, EmailMultiAlternatives
from django.db import IntegrityError, transaction
from django.db.models import Q, F, Sum, Count
from django.db.models.functions import Coalesce, Least
from django.shortcuts import render, redirect, get_object_or_404
from django.template import Template, Context
from django.template.loader import render_to_string
from django.urls import reverse, reverse_lazy
from django.utils.dateparse import parse_date
from django.utils.text import slugify

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from django.core.paginator import Paginator
from django.contrib.admin.views.decorators import staff_member_required
from operator import itemgetter
import csv
import calendar
import random
import string
from io import TextIOWrapper
from django.http import HttpResponse, FileResponse, HttpResponseForbidden, JsonResponse, HttpResponseNotFound
from django.utils import timezone
from django.contrib import messages
from django.views.decorators.http import require_POST, require_http_methods
from .models import (
    Attraction,
    VisitSlot,
    Booking,
    Profile,
    TicketDraw,
    TicketDrawBooking,
    TicketDrawVisitSlot,
    AttractionSuggestion,
    AttractionWaitlistEntry,
    DiscountCode,
    EmailTemplate, BookingTicket,
    EmailTemplate,
    FeedbackEmailTemplate,
)

from .forms import (
    BookingForm,
    AttractionCreateForm,
    TicketDrawCreateForm,
    EmailAuthenticationForm,
    FeedbackEmailTemplateForm
)

from .forms_discount_codes import DiscountCodeForm
from .forms_suggestions import AttractionSuggestionForm

User = get_user_model()

# Business rule: max 3 bookings per calendar year (regular) and 3 per year (weekly_event) unless overridden elsewhere.
MAX_ATTRACTIONS_PER_YEAR = 3


# -----------------------------
# Small helpers / utilities
# -----------------------------
def _call_is_open(obj, now):
    """
    Supports both signatures:
      - obj.is_open()
      - obj.is_open(now)
    """
    if not hasattr(obj, "is_open"):
        return None
    try:
        return obj.is_open(now)
    except TypeError:
        return obj.is_open()


def calculate_remaining_allowance(user, attraction_type="regular"):
    """
    attraction_type:
      - 'regular'      -> Booking against Attraction
      - 'weekly_event' -> TicketDrawBooking against TicketDraw
    """
    year = timezone.now().year

    if attraction_type == "regular":
        used = Booking.objects.filter(
            user=user,
            cancelled=False,
            created_at__year=year,
            attraction__attraction_type="regular",
        ).count()
        return max(0, MAX_ATTRACTIONS_PER_YEAR - used)

    if attraction_type == "weekly_event":
        used = TicketDrawBooking.objects.filter(
            user=user,
            cancelled=False,
            created_at__year=year,
            ticket_draw__attraction_type="weekly_event",
        ).count()
        return max(0, MAX_ATTRACTIONS_PER_YEAR - used)

    return 0


def add_events(objects, events_by_day, start, end, event_type):
    """Add booking_open / booking_close events for calendar display."""
    for obj in objects:
        date_value = getattr(obj, "date", None)
        if not date_value:
            continue
        event_date = date_value
        if start <= event_date <= end:
            events_by_day.setdefault(event_date.day, []).append(
                {"object": obj, "event_type": event_type}
            )


def get_calendar(year=None, month=None):
    """
    Build calendar data for dashboard/calendar template rendering.
    Returns a dict (NOT a rendered response), so it can be merged into context via **calendar_data.
    """
    today = timezone.localdate()
    year = int(year or today.year)
    month = int(month or today.month)

    # prev / next nav
    if month == 1:
        prev_month, prev_year = 12, year - 1
    else:
        prev_month, prev_year = month - 1, year

    if month == 12:
        next_month, next_year = 1, year + 1
    else:
        next_month, next_year = month + 1, year

    cal = calendar.Calendar(firstweekday=0)
    month_days = list(cal.itermonthdates(year, month))
    start, end = month_days[0], month_days[-1]

    events_by_day = {}
    add_events(VisitSlot.objects.all(), events_by_day, start, end, "attraction")
    add_events(TicketDrawVisitSlot.objects.all(), events_by_day, start, end, "ticket_draw")

    weeks = []
    for i in range(0, len(month_days), 7):
        week = month_days[i : i + 7]
        week_info = []
        for day in week:
            day_events = events_by_day.get(day.day, []) if day.month == month else []
            week_info.append({"date": day, "events": day_events})
        weeks.append(week_info)

    return {
        "year": year,
        "month": month,
        "month_name": calendar.month_name[month],
        "weeks": weeks,
        "today": today,
        "prev_year": prev_year,
        "prev_month": prev_month,
        "next_year": next_year,
        "next_month": next_month,
    }


def assign_next_winner(draw: TicketDraw):
    """
    If current winner entry is cancelled/missing or deadline passed, pick a new winner from active entries.
    If no active entries then clear winner.
    """
    if getattr(draw, "winner_booking", None):
        winner = draw.winner_booking

        if winner.is_accepted:
            return

        if winner.cancelled:
            draw.winner_booking = None
            draw.winner_selected_at = None

        # Check if winner has exceeded 72 hour deadline
        elif draw.winner_selected_at:
            deadline = draw.winner_selected_at + datetime.timedelta(hours=72)
            if timezone.now() > deadline:
                winner.cancelled = True
                winner.save(update_fields=["cancelled"])

                # Restore slot capacity
                TicketDrawVisitSlot.objects.filter(pk=winner.slot_id).update(
                    remaining=F("remaining") + winner.num_tickets
                )

                draw.winner_booking = None
                draw.winner_selected_at = None
            else:
                return

    # Get all active entries (not cancelled, not accepted)
    entries = list(
        TicketDrawBooking.objects.filter(ticket_draw=draw, cancelled=False, is_accepted=False).select_related("user",
                                                                                                              "ticket_draw",
                                                                                                              "slot")
    )

    if not entries:
        draw.winner_booking = None
        draw.winner_selected_at = None
    else:
        draw.winner_booking = random.choice(entries)
        draw.winner_selected_at = timezone.now()

        send_draw_booking_email_redraw_winner(draw.winner_booking)

    draw.save(update_fields=["winner_booking", "winner_selected_at"])


# -----------------------------
# Auth
# -----------------------------
class CustomLoginView(LoginView):
    template_name = "fergusonbequest/login.html"
    authentication_form = EmailAuthenticationForm

    def get_success_url(self):
        return reverse_lazy("home")


class RegistrationForm(forms.ModelForm):
    """Registration form for User. Unique email, password confirmation, auto-username."""

    password = forms.CharField(widget=forms.PasswordInput, label="Password")
    password_confirm = forms.CharField(widget=forms.PasswordInput, label="Confirm Password")

    class Meta:
        model = User
        fields = ("first_name", "last_name", "email")

    def clean(self):
        cleaned = super().clean()
        p = cleaned.get("password")
        pc = cleaned.get("password_confirm")
        if p and pc and p != pc:
            self.add_error("password_confirm", "Passwords do not match.")
        return cleaned

    def clean_email(self):
        email = (self.cleaned_data.get("email") or "").strip().lower()
        if User.objects.filter(email__iexact=email).exists():
            raise forms.ValidationError("A user with that email already exists.")
        return email

    def save(self, commit=True):
        user = super().save(commit=False)

        first = (self.cleaned_data.get("first_name") or "").strip()
        last = (self.cleaned_data.get("last_name") or "").strip()

        if first or last:
            base = f"{first.capitalize()}{last.capitalize()}"
        else:
            email = self.cleaned_data.get("email") or ""
            base = email.split("@")[0] if "@" in email else "user"

        base = base or "user"
        username = base
        counter = 0
        while User.objects.filter(username=username).exists():
            counter += 1
            username = f"{base}{counter}"

        user.username = username
        user.email = (self.cleaned_data.get("email") or "").strip().lower()
        user.set_password(self.cleaned_data.get("password"))

        if commit:
            user.save()
        return user


def register_view(request):
    if request.method == "POST":
        form = RegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, "Registration successful.")
            return redirect("home")
    else:
        form = RegistrationForm()

    return render(request, "fergusonbequest/register.html", {"form": form})


@require_POST
def logout_view(request):
    logout(request)
    messages.info(request, "You have been logged out.")
    return redirect("home")


# -----------------------------
# Public / Home / Dashboard / Calendar / Terms
# -----------------------------
def home(request):
    attractions_qs = Attraction.objects.all().order_by("name")[:4]

    featured_attractions = [
        {
            "title": attr.name,
            "subtitle": (attr.description[:100] if attr.description else (attr.location or "Book now to visit")),
            "image": (attr.image.name if getattr(attr, "image", None) else "fergusonbequest/img/placeholder.jpg"),
            "id": attr.id,
            "url": f"/attraction/{attr.id}/book/",
        }
        for attr in attractions_qs
    ]

    # fallback if DB empty (helps tests/first run)
    if not featured_attractions:
        featured_attractions = [
            {"title": "Blair Drummond Safari Park", "subtitle": "Safari and adventure park.",
             "image": "fergusonbequest/img/blair_drumond.jpg", "id": None, "url": "/attractions/"},
            {"title": "Glasgow Clan Ice Hockey", "subtitle": "The city's professional hockey team.",
             "image": "fergusonbequest/img/glasgow_clan.jpg", "id": None, "url": "/attractions/"},
            {"title": "Edinburgh Zoo", "subtitle": "Scotland's most famous zoo.",
             "image": "fergusonbequest/img/edinburgh_zoo.jpg", "id": None, "url": "/attractions/"},
            {"title": "Ghostbusters Screening", "subtitle": "Who you gonna call?",
             "image": "fergusonbequest/img/ghostbusters.jpg", "id": None, "url": "/attractions/"},
        ]

    if request.user.is_authenticated:
        return render(
            request,
            "fergusonbequest/home_logged_in.html",
            {"featured_attractions": featured_attractions},
        )

    return render(request, "fergusonbequest/home.html", {"featured_attractions": featured_attractions})


@login_required
def dashboard_view(request, year=None, month=None):
    attractions_qs = Attraction.objects.all().order_by("name")[:4]

    featured_attractions = [
        {
            "title": attr.name,
            "subtitle": (attr.description[:100] if attr.description else (attr.location or "Book now to visit")),
            "image": (attr.image.name if getattr(attr, "image", None) else "fergusonbequest/img/placeholder.jpg"),
            "id": attr.id,
            "url": f"/attraction/{attr.id}/book/",
        }
        for attr in attractions_qs
    ]

    if not featured_attractions:
        featured_attractions = [
            {"title": "Blair Drummond Safari Park", "subtitle": "Safari and adventure park.",
             "image": "fergusonbequest/img/blair_drumond.jpg", "id": None, "url": "/attractions/"},
            {"title": "Glasgow Clan Ice Hockey", "subtitle": "The city's professional hockey team.",
             "image": "fergusonbequest/img/glasgow_clan.jpg", "id": None, "url": "/attractions/"},
            {"title": "Edinburgh Zoo", "subtitle": "Scotland's most famous zoo.",
             "image": "fergusonbequest/img/edinburgh_zoo.jpg", "id": None, "url": "/attractions/"},
            {"title": "Ghostbusters Screening", "subtitle": "Who you gonna call?",
             "image": "fergusonbequest/img/ghostbusters.jpg", "id": None, "url": "/attractions/"},
        ]

    calendar_data = get_calendar(year, month)
    return render(
        request,
        "fergusonbequest/dashboard.html",
        {"featured_attractions": featured_attractions, **calendar_data},
    )


def calendar_view(request, year=None, month=None):
    context = get_calendar(year, month)
    return render(request, "fergusonbequest/calendar.html", context)


def terms(request):
    from .models import TermsAndConditions
    t_and_c = TermsAndConditions.get()
    return render(request, "fergusonbequest/terms.html", {'t_and_c': t_and_c})


# -----------------------------
# Attractions (user)
# -----------------------------
def attractions_view(request):
    q = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "").strip()

    date_str = (request.GET.get("date") or "").strip()
    d = parse_date(date_str) if date_str else None

    location = (request.GET.get("location") or "").strip()
    sort = (request.GET.get("sort") or "name").strip()
    type_filter = (request.GET.get("type") or "").strip()

    today = timezone.now().date()

    attractions = (
        Attraction.objects.annotate(
            future_slots_count=Count("slots", filter=Q(slots__date__gte=today), distinct=True),
            tickets_left_total=Coalesce(Sum("slots__remaining", filter=Q(slots__date__gte=today)), 0),
        )
        .filter(future_slots_count__gt=0)
        .distinct()
    )

    if q:
        attractions = attractions.filter(Q(name__icontains=q) | Q(location__icontains=q))

    if status == "available":
        attractions = attractions.filter(tickets_left_total__gt=0)
    elif status == "soldout":
        attractions = attractions.filter(tickets_left_total=0)

    if location:
        attractions = attractions.filter(location__iexact=location)

    if d:
        attractions = attractions.filter(slots__date__gte=d).distinct()

    if sort == "tickets":
        attractions = attractions.order_by("-tickets_left_total", "name")
    else:
        attractions = attractions.order_by("name")

    if type_filter:
        attractions = attractions.filter(attraction_type=type_filter)

    locations = Attraction.objects.values_list("location", flat=True).distinct().order_by("location")

    for a in attractions:
        future_slots = list(
            VisitSlot.objects.filter(
                attraction=a,
                date__gte=today,
            ).order_by("date", "time")
        )

        a.sold_out_slots = [s for s in future_slots if s.remaining == 0]
        a.sold_out_slot_count = len(a.sold_out_slots)

    return render(
        request,
        "fergusonbequest/attractions.html",
        {
            "attractions": attractions,
            "q": q,
            "status": status,
            "date": date_str,
            "location_filter": location,
            "sort": sort,
            "locations": locations,
            "types": ["regular", "weekly_event"],
            "type_filter": type_filter,
        },
    )


def attraction(request, pk):
    attraction_obj = get_object_or_404(Attraction, pk=pk)

    available_slots = VisitSlot.objects.filter(
        attraction=attraction_obj,
        date__gte=timezone.now().date(),
    ).order_by("date", "time")

    bookable_slots = available_slots.filter(remaining__gt=0)
    has_bookable_slots = bookable_slots.exists()
    is_sold_out = available_slots.exists() and not has_bookable_slots

    remaining_allowance = 0
    waitlisted_slot_ids = set()

    if request.user.is_authenticated:
        remaining_allowance = calculate_remaining_allowance(
            request.user,
            attraction_obj.attraction_type
        )

        waitlisted_slot_ids = set(
            AttractionWaitlistEntry.objects.filter(
                user=request.user,
                cancelled=False,
                slot__in=available_slots,
            ).values_list("slot_id", flat=True)
        )

    return render(
        request,
        "fergusonbequest/attraction.html",
        {
            "attraction": attraction_obj,
            "available_slots": available_slots,
            "bookable_slots": bookable_slots,
            "has_bookable_slots": has_bookable_slots,
            "remaining_allowance": remaining_allowance,
            "is_sold_out": is_sold_out,
            "waitlisted_slot_ids": waitlisted_slot_ids,
        },
    )


@login_required
def booking_view(request, attraction_pk):
    attraction_obj = get_object_or_404(Attraction, pk=attraction_pk)
    user = request.user

    available_slots = VisitSlot.objects.filter(
        attraction=attraction_obj,
        date__gte=timezone.now().date(),
    ).order_by("date", "time")

    booking_summary = {"price": "Free"}
    remaining_allowance = calculate_remaining_allowance(user, "regular")

    def redirect_to_history_with(msg):
        messages.error(request, msg)
        return redirect("booking_history")

    # GET
    if request.method == "GET":
        if remaining_allowance <= 0:
            return redirect_to_history_with(
                f"You have reached your yearly limit of {MAX_ATTRACTIONS_PER_YEAR} attractions. "
                "Please cancel/delete an existing booking in Booking History before booking again."
            )

        slot_id = request.GET.get("slot")
        if slot_id and not available_slots.filter(pk=slot_id).exists():
            return redirect_to_history_with("Invalid slot selected for this attraction.")

        if slot_id:
            already = Booking.objects.filter(user=user, slot_id=slot_id, cancelled=False).exists()
            if already:
                return redirect_to_history_with(
                    "Oops — you already have a booking for this time slot. "
                    "Please cancel it in Booking History before booking again."
                )

        form = BookingForm(attraction=attraction_obj, initial={"slot": slot_id} if slot_id else None)
        return render(
            request,
            "fergusonbequest/booking_page.html",
            {
                "attraction": attraction_obj,
                "available_slots": available_slots,
                "form": form,
                "booking_summary": booking_summary,
                "remaining_allowance": remaining_allowance,
                "max_allowance": MAX_ATTRACTIONS_PER_YEAR,
                "now": timezone.now(),
            },
        )

    # POST
    form = BookingForm(request.POST, attraction=attraction_obj)
    if not form.is_valid():
        return render(
            request,
            "fergusonbequest/booking_page.html",
            {
                "attraction": attraction_obj,
                "available_slots": available_slots,
                "form": form,
                "booking_summary": booking_summary,
                "remaining_allowance": remaining_allowance,
                "max_allowance": MAX_ATTRACTIONS_PER_YEAR,
                "now": timezone.now(),
            },
        )

    # Re-check allowance (safety)
    remaining_allowance = calculate_remaining_allowance(user, "regular")
    if remaining_allowance <= 0:
        return redirect_to_history_with(
            f"You have reached your yearly limit of {MAX_ATTRACTIONS_PER_YEAR} attractions. "
            "Please cancel/delete an existing booking in Booking History before booking again."
        )

    booking = form.save(commit=False)
    booking.user = user
    booking.attraction = attraction_obj
    booking.full_name = f"{user.first_name} {user.last_name}".strip()
    booking.email = user.email

    visit_datetime = datetime.combine(booking.slot.date, datetime.min.time())

    # Make it timezone-aware if needed
    if timezone.is_aware(timezone.now()):
        import pytz
        visit_datetime = timezone.make_aware(visit_datetime)
    
    # Set cancellation deadline to 3 days before the visit
    booking.cancel_deadline = visit_datetime - timedelta(days=3)

    if booking.slot.attraction_id != attraction_obj.id:
        return redirect_to_history_with("Invalid slot selected for this attraction.")

    already = Booking.objects.filter(user=user, slot=booking.slot, cancelled=False).exists()
    if already:
        return redirect_to_history_with(
            "Oops — you already have a booking for this time slot. "
            "Please cancel it in Booking History before booking again."
        )

    try:
        with transaction.atomic():
            slot = VisitSlot.objects.select_for_update().get(pk=booking.slot.pk)
            if slot.remaining < booking.num_tickets:
                form.add_error("slot", "Not enough tickets remaining for this slot.")
                return render(
                    request,
                    "fergusonbequest/booking_page.html",
                    {
                        "attraction": attraction_obj,
                        "available_slots": available_slots,
                        "form": form,
                        "booking_summary": booking_summary,
                        "remaining_allowance": remaining_allowance,
                        "max_allowance": MAX_ATTRACTIONS_PER_YEAR,
                        "now": timezone.now(),
                    },
                )

            booking.save()
            VisitSlot.objects.filter(pk=slot.pk).update(remaining=F("remaining") - booking.num_tickets)

        messages.success(request, "Booking confirmed!")
        return redirect("booking_history")

    except IntegrityError:
        return redirect_to_history_with(
            "Duplicate booking detected for this time slot. Please cancel it in Booking History before booking again."
        )



@login_required
def booking_history(request):
    """Show list of bookings for the logged in user, including both attractions and ticket draws."""
    user = request.user

    remaining_allowance = calculate_remaining_allowance(user, 'regular')

    # Base querysets for both types
    bookings = Booking.objects.filter(user=user).select_related('slot', 'attraction')
    draw_bookings = TicketDrawBooking.objects.filter(user=user).select_related('slot', 'ticket_draw')

    # Parse GET params for filters
    when = request.GET.get('when')  # all|future|past|cancelled
    status = request.GET.get('status')  # all|active|cancelled
    venue = request.GET.get('venue')
    q = request.GET.get('q')
    start = request.GET.get('start')
    end = request.GET.get('end')
    sort = request.GET.get('sort', 'created_at')
    booking_type = (request.GET.get("type") or "").lower()  # attraction|draw|both

    today = timezone.now().date()

    # Apply status filters
    if status == 'cancelled':
        bookings = bookings.filter(cancelled=True)
        draw_bookings = draw_bookings.filter(cancelled=True)
    elif status == 'active':
        bookings = bookings.filter(cancelled=False)
        draw_bookings = draw_bookings.filter(cancelled=False)

    # Apply when filters
    if when == 'future':
        bookings = bookings.filter(slot__date__gte=today)
        draw_bookings = draw_bookings.filter(slot__date__gte=today)
    elif when == 'past':
        bookings = bookings.filter(slot__date__lt=today)
        draw_bookings = draw_bookings.filter(slot__date__lt=today)
    elif when == 'cancelled':
        bookings = bookings.filter(cancelled=True)
        draw_bookings = draw_bookings.filter(cancelled=True)

    # Venue filter
    if venue:
        if venue.isdigit():
            bookings = bookings.filter(attraction__id=venue)
            draw_bookings = draw_bookings.filter(ticket_draw__id=venue)
        else:
            bookings = bookings.filter(attraction__slug__icontains=venue)
            draw_bookings = draw_bookings.filter(ticket_draw__slug__icontains=venue)

    # Apply booking type filter
    if booking_type == "attraction":
        draw_bookings = draw_bookings.none()
    elif booking_type == "draw":
        bookings = bookings.none()

    if q:
        bookings = bookings.filter(
            Q(attraction__name__icontains=q) |
            Q(id__icontains=q) |
            Q(email__icontains=q) |
            Q(ticket_code__icontains=q)
        )
        draw_bookings = draw_bookings.filter(
            Q(ticket_draw__name__icontains=q) |
            Q(id__icontains=q) |
            Q(email__icontains=q) |
            Q(converted_booking__ticket_code__icontains=q)
        )

    sd = parse_date(start) if start else None
    ed = parse_date(end) if end else None
    if sd:
        bookings = bookings.filter(slot__date__gte=sd)
        draw_bookings = draw_bookings.filter(slot__date__gte=sd)
    if ed:
        bookings = bookings.filter(slot__date__lte=ed)
        draw_bookings = draw_bookings.filter(slot__date__lte=ed)

    # Split into past/future and combine
    future_bookings = list(bookings.filter(slot__date__gte=today))
    future_draws = list(draw_bookings.filter(slot__date__gte=today))

    past_bookings = list(bookings.filter(slot__date__lt=today))
    past_draws = list(draw_bookings.filter(slot__date__lt=today))

    # Ticket release helpers
    def apply_ticket_release_flags(obj, *, is_draw: bool):
        """
        Ticket visibility rules:
        - If obj.ticket_visible_at is set: ticket visible when timezone.now() >= ticket_visible_at
        - Else: visible when days_until_visit <= ticket_release_days
        """
        now_dt = timezone.now()
        today_local = now_dt.date()

        visit_date = getattr(getattr(obj, "slot", None), "date", None)

        # Pull release-days from the related venue/draw if present, else default 0.
        src = getattr(obj, "ticket_draw", None) if is_draw else getattr(obj, "attraction", None)
        release_days = int(getattr(src, "ticket_release_days", 0) or 0)

        # Only meaningful if there IS a ticket configured/assigned.
        ticket_obj = obj.converted_booking if is_draw and getattr(obj, "converted_booking", None) else obj
        has_ticket = bool(
            getattr(ticket_obj, "ticket_type", None)
            or getattr(ticket_obj, "ticket_code", None)
            or getattr(ticket_obj, "generic_booking_code", None)
            or getattr(ticket_obj, "ticket_instructions", None)
            or getattr(ticket_obj, "box_office_notes", None)
            or (hasattr(ticket_obj, "tickets") and ticket_obj.tickets.exists())
        )

        # Defaults
        can_view = False
        days_until = None
        days_to_release = None

        # Scheduled visibility takes priority
        visible_at = getattr(obj, "ticket_visible_at", None)
        if visible_at:
            can_view = has_ticket and (now_dt >= visible_at)
            if not can_view:
                delta = visible_at - now_dt
                days_to_release = max(0, delta.days)

        # Fallback to "X days before visit"
        else:
            if visit_date:
                days_until = (visit_date - today_local).days

            if has_ticket and days_until is not None:
                can_view = (days_until <= release_days)

            if days_until is not None:
                days_to_release = max(0, days_until - release_days)

        obj.ticket_release_days = release_days
        obj.days_until_visit = days_until
        obj.days_to_release = days_to_release
        obj.can_view_ticket = can_view

    # flags + apply the helper
    for b in future_bookings + past_bookings:
        b.booking_type = "attraction"
        b.is_draw = False
        apply_ticket_release_flags(b, is_draw=False)

    for d in future_draws + past_draws:
        d.booking_type = "draw"
        d.is_draw = True
        apply_ticket_release_flags(d, is_draw=True)

    # Combine and sort
    if sort == "slot_date":

        def sort_key(item):
            return (item.slot.date, item.created_at)

        reverse_sort = False

    else:

        def sort_key(item):
            return item.created_at

        reverse_sort = True

    future_all = sorted(future_bookings + future_draws, key=sort_key, reverse=reverse_sort)
    past_all = sorted(past_bookings + past_draws, key=sort_key, reverse=reverse_sort)

    return render(
        request,
        "fergusonbequest/booking_history.html",
        {
            "future_bookings": future_all,
            "past_bookings": past_all,
            "when": when,
            "status": status,
            "venue": venue,
            "q": q,
            "start": start,
            "end": end,
            "sort": sort,
            "now": timezone.now(),
            "remaining_allowance": remaining_allowance,
            "max_allowance": MAX_ATTRACTIONS_PER_YEAR,
            "today": today,
            "user_ticket_url_template": reverse("user_ticket_view", args=[999999]),
        },
    )

@staff_member_required
def manage_feedback_email(request):
    """
    Allow admins to edit the feedback email template through a simple web form.
    """
    template = FeedbackEmailTemplate.get_template()

    feedback_email_template = (
        EmailTemplate.objects.filter(type="feedback", is_default=True).first()
        or EmailTemplate.objects.filter(type="feedback").first()
    )

    if feedback_email_template:
        changed_fields = []
        if template.subject != feedback_email_template.subject:
            template.subject = feedback_email_template.subject
            changed_fields.append("subject")
        if template.body != feedback_email_template.body:
            template.body = feedback_email_template.body
            changed_fields.append("body")
        if changed_fields:
            template.save(update_fields=changed_fields)
    
    if request.method == 'POST':
        form = FeedbackEmailTemplateForm(request.POST, instance=template)
        if form.is_valid():
            saved_template = form.save()

            email_template_defaults = {
                "name": "Feedback Template",
                "subject": saved_template.subject,
                "body": saved_template.body,
                "is_default": True,
            }

            feedback_email_template = (
                EmailTemplate.objects.filter(type="feedback", is_default=True).first()
                or EmailTemplate.objects.filter(type="feedback").first()
            )

            if feedback_email_template:
                feedback_email_template.subject = saved_template.subject
                feedback_email_template.body = saved_template.body
                if not feedback_email_template.is_default:
                    EmailTemplate.objects.filter(type="feedback").update(is_default=False)
                    feedback_email_template.is_default = True
                feedback_email_template.save()
            else:
                EmailTemplate.objects.filter(type="feedback").update(is_default=False)
                EmailTemplate.objects.create(type="feedback", **email_template_defaults)

            messages.success(request, 'Feedback email settings saved successfully!')
            return redirect('manage_feedback_email')
    else:
        form = FeedbackEmailTemplateForm(instance=template)
    
    return render(request, 'fergusonbequest/manage_feedback_email.html', {
        'form': form,
        'template': template,
    })


@staff_member_required
@require_POST
def trigger_feedback_emails(request):
    """Allow staff to manually trigger the feedback email send from the web UI."""
    from .scheduler import send_scheduled_feedback_emails
    try:
        send_scheduled_feedback_emails()
        messages.success(request, 'Feedback emails sent successfully.')
    except Exception as e:
        messages.error(request, f'Error sending feedback emails: {e}')
    return redirect('manage_feedback_email')


@require_POST
@login_required
def cancel_booking(request, pk):
    """Cancel a future booking and restore slot capacity."""
    booking = get_object_or_404(Booking, pk=pk)

    if not (request.user == booking.user or request.user.is_superuser):
        return redirect("booking_history")

    if booking.slot.date < timezone.now().date():
        return redirect("booking_history")

    reassigned_booking = None

    with transaction.atomic():
        b = Booking.objects.select_for_update().select_related("slot", "attraction", "user").get(pk=booking.pk)

        if not b.cancelled:
            b.cancelled = True
            b.save(update_fields=["cancelled"])

            VisitSlot.objects.filter(pk=b.slot.pk).update(
                remaining=Least(F("remaining") + b.num_tickets, F("capacity"))
            )

            # reload the slot after increment
            slot = VisitSlot.objects.select_for_update().get(pk=b.slot.pk)

            # Only try reassignment for attraction bookings
            reassigned_booking = reassign_cancelled_attraction_booking(slot, b.attraction)

            send_attraction_booking_email_cancellation(b)

            current_year = timezone.now().year
            active_yearly_count = Booking.objects.filter(
                user=b.user, cancelled=False, created_at__year=current_year
            ).count()
            remaining = max(0, MAX_ATTRACTIONS_PER_YEAR - active_yearly_count)

            if reassigned_booking:
                messages.success(
                    request,
                    f"Booking cancelled. The freed slot was reassigned to a user on the waiting list. "
                    f"You now have {remaining}/{MAX_ATTRACTIONS_PER_YEAR} bookings remaining for this year.",
                )
            else:
                messages.success(
                    request,
                    f"Booking cancelled. You now have {remaining}/{MAX_ATTRACTIONS_PER_YEAR} bookings remaining for this year.",
                )

    return redirect("booking_history")


# -----------------------------
# Ticket draws (user)
# -----------------------------
@login_required
def ticket_draws_view(request):
    draws = TicketDraw.objects.all().order_by("name")

    selected = None
    date_range = None

    selected_id = request.GET.get("draw")
    if selected_id:
        selected = get_object_or_404(TicketDraw, pk=selected_id)

        slots = TicketDrawVisitSlot.objects.filter(ticket_draw=selected).order_by("date")
        if slots.exists():
            first_date = slots.first().date
            last_date = slots.last().date
            if first_date == last_date:
                date_range = first_date.strftime("%d/%m/%Y")
            else:
                date_range = f"{first_date.strftime('%d/%m/%Y')} – {last_date.strftime('%d/%m/%Y')}"
        else:
            start_date = getattr(selected, "start_date", None)
            end_date = getattr(selected, "end_date", None)
            draw_date = getattr(selected, "draw_date", None)
            if draw_date:
                try:
                    date_range = draw_date.strftime("%d/%m/%Y")
                except Exception:
                    date_range = str(draw_date)
            elif start_date and end_date:
                try:
                    date_range = (
                        start_date.strftime("%d/%m/%Y")
                        if start_date == end_date
                        else f"{start_date.strftime('%d/%m/%Y')} – {end_date.strftime('%d/%m/%Y')}"
                    )
                except Exception:
                    date_range = f"{start_date} – {end_date}"

    return render(
        request,
        "fergusonbequest/ticket_draws.html",
        {"draws": draws, "selected": selected, "date_range": date_range},
    )


@login_required
def ticket_draw_detail(request, slug):
    draw = get_object_or_404(TicketDraw, slug=slug)

    existing_entries = TicketDrawBooking.objects.filter(
        user=request.user,
        ticket_draw=draw,
        cancelled=False,
    ).count()

    remaining_allowance = calculate_remaining_allowance(
        request.user, getattr(draw, "attraction_type", "weekly_event")
    )
    draw_limit = getattr(draw, "per_year_limit", MAX_ATTRACTIONS_PER_YEAR)
    draw_specific_remaining = max(0, draw_limit - existing_entries)
    remaining_allowance = min(remaining_allowance, draw_specific_remaining)

    if request.method == "POST":
        if existing_entries > 0:
            messages.error(
                request,
                f"You already have an active entry for {draw.name}. "
                "You must cancel your existing entry before booking a different date or adding tickets.",
            )
            return redirect("draw_waiting_list")

        num_tickets = int(request.POST.get("num_tickets", 1))
        if num_tickets > remaining_allowance:
            messages.error(request, f"Max limit reached. You can only choose up to {remaining_allowance} more tickets.")
            return redirect("ticket_draw_detail", slug=slug)

        slot_id = request.POST.get("slot_id")
        if not slot_id:
            messages.error(request, "No available dates for this draw. Please check back later.")
            return redirect("ticket_draw_detail", slug=slug)

        slot = TicketDrawVisitSlot.objects.filter(pk=slot_id, ticket_draw=draw).first()
        if slot is None:
            messages.error(request, "Selected date is no longer available. Please choose another date.")
            return redirect("ticket_draw_detail", slug=slug)

        # check draw open
        if not _call_is_open(draw, timezone.now()):
            messages.error(request, "This draw is currently closed.")
            return redirect("ticket_draw_detail", slug=slug)

        if slot.remaining < num_tickets:
            messages.error(request, "Not enough availability for that date.")
            return redirect("ticket_draw_detail", slug=slug)

        with transaction.atomic():
            draw_booking = TicketDrawBooking.objects.create(
                user=request.user,
                ticket_draw=draw,
                slot=slot,
                num_tickets=num_tickets,
                full_name=f"{request.user.first_name} {request.user.last_name}",
                email=request.user.email,
                agreed_terms=True,
            )

            TicketDrawVisitSlot.objects.filter(pk=slot.pk).update(
                remaining=F("remaining") - num_tickets
            )

        messages.success(request, "Successfully entered draw!")
        send_draw_booking_email_confirmation(draw_booking)
        return redirect("draw_waiting_list")

    slots = TicketDrawVisitSlot.objects.filter(
        ticket_draw=draw,
        date__gte=timezone.now().date(),
    ).order_by("date", "time")

    return render(
        request,
        "fergusonbequest/ticket_draw_detail.html",
        {"draw": draw, "slots": slots, "remaining_allowance": remaining_allowance},
    )


@login_required
def draw_waiting_list(request):
    """Pending draw entries for the user; highlights winner needing action."""
    user = request.user
    now = timezone.now()

    entries = (
        TicketDrawBooking.objects.filter(user=user, cancelled=False, is_accepted=False)
        .select_related("ticket_draw", "slot")
        .order_by("-created_at")
    )

    for e in entries:
        is_winner = getattr(e.ticket_draw, "winner_booking_id", None) == e.id
        if is_winner and getattr(e, "is_accepted", False):
            e.ui_status = "Accepted"
        elif is_winner:
            e.ui_status = "Winner (Action required)"
        else:
            e.ui_status = "Waiting for draw"

        close = getattr(e.ticket_draw, "booking_close", None)
        e.can_cancel = (not getattr(e, "is_accepted", False)) and (not is_winner) and bool(close and now <= close)

    return render(request, "fergusonbequest/draw_waiting_list.html", {"entries": entries, "now": now})


# Compatibility alias (older routes may still point here)
@login_required
def waiting_list(request):
    return redirect("draw_waiting_list")


@require_POST
@login_required
def cancel_ticket_draw_entry(request, pk):
    booking = get_object_or_404(TicketDrawBooking, pk=pk, user=request.user)

    if getattr(booking, "is_accepted", False):
        messages.error(request, "You can't cancel an accepted draw win. Please contact support.")
        return redirect("draw_waiting_list")

    with transaction.atomic():
        if not booking.cancelled:
            booking.cancelled = True
            booking.save(update_fields=["cancelled"])

            TicketDrawVisitSlot.objects.filter(pk=booking.slot_id).update(
                remaining=F("remaining") + booking.num_tickets
            )

            draw = booking.ticket_draw
            if getattr(draw, "winner_booking_id", None) == booking.id:
                draw.winner_booking = None
                draw.winner_selected_at = None
                draw.save(update_fields=["winner_booking", "winner_selected_at"])
                assign_next_winner(draw)

            messages.success(request, f"Entry for {draw.name} cancelled.")
        else:
            messages.info(request, "This entry was already cancelled.")

    return redirect(request.META.get("HTTP_REFERER", "draw_waiting_list"))


@transaction.atomic
def accept_draw_booking(draw_booking):
    if draw_booking.converted_booking_id:
        draw_booking.is_accepted = True
        draw_booking.save(update_fields=["is_accepted"])
        return draw_booking.converted_booking

    draw = draw_booking.ticket_draw
    draw_slot = draw_booking.slot

    attraction, _ = Attraction.objects.get_or_create(
        slug=slugify(draw.name),
        defaults={
            "name": draw.name,
            "location": draw.location,
            "per_year_limit": draw.per_year_limit,
            "booking_open": draw.booking_open,
            "booking_close": draw.booking_close,
        },
    )

    visit_slot, _ = VisitSlot.objects.get_or_create(
        attraction=attraction,
        date=draw_slot.date,
        time=draw_slot.time,
        defaults={
            "capacity": draw_slot.capacity,
            "remaining": max(draw_slot.remaining, 0),
        },
    )

    booking = Booking.objects.create(
        user=draw_booking.user,
        attraction=attraction,
        slot=visit_slot,
        full_name=draw_booking.full_name,
        email=draw_booking.email,
        num_tickets=draw_booking.num_tickets,
        agreed_terms=True,
        cancelled=False,
    )

    draw_booking.is_accepted = True
    draw_booking.converted_booking = booking
    draw_booking.save(update_fields=["is_accepted", "converted_booking"])

    return booking

@login_required
#@require_POST doesnt work with email links if require post
def accept_draw_win(request, pk):
    draw_booking = get_object_or_404(TicketDrawBooking, pk=pk, user=request.user)
    draw = draw_booking.ticket_draw

    if draw.winner_booking_id != draw_booking.id:
        messages.error(request, "You are not the current winner of this draw.")
        return redirect("booking_history")

    if draw_booking.is_accepted and draw_booking.converted_booking_id:
        messages.info(request, "You have already accepted these tickets.")
        return redirect("booking_history")

    with transaction.atomic():
        accept_draw_booking(draw_booking)

    messages.success(request, f"You have accepted your tickets for {draw.name}.")
    return redirect("booking_history")


@login_required
def decline_draw_win(request, pk):
    booking = get_object_or_404(TicketDrawBooking, pk=pk, user=request.user)
    draw = booking.ticket_draw

    if getattr(draw, "winner_booking_id", None) != booking.id:
        messages.error(request, "Invalid request.")
        return redirect("draw_waiting_list")

    if booking.is_accepted:
        messages.error(request, "You have already accepted these tickets and cannot decline them now.")
        return redirect("booking_history")

    with transaction.atomic():
        if not booking.cancelled:
            booking.cancelled = True
            booking.save(update_fields=["cancelled"])

            TicketDrawVisitSlot.objects.filter(pk=booking.slot_id).update(
                remaining=F("remaining") + booking.num_tickets)
            # Send decline email
            send_draw_booking_email_cancellation(booking)

        draw.winner_booking = None
        draw.winner_selected_at = None
        draw.save(update_fields=["winner_booking", "winner_selected_at"])

        assign_next_winner(draw)

    messages.info(request, f"You have declined the tickets for {draw.name}. They will be offered to someone else.")
    return redirect("draw_waiting_list")

# -----------------------------
# Waiting list for attractions
# -----------------------------
@login_required
def waiting_listattraction(request):
    """Attraction waiting list"""
    user = request.user

    attraction_waitlist_entries = (
        AttractionWaitlistEntry.objects
        .filter(user=user, cancelled=False)
        .select_related("attraction", "slot")
        .order_by("-created_at")
    )

    return render(
        request,
        "fergusonbequest/waiting_listattraction.html",
        {"attraction_waitlist_entries": attraction_waitlist_entries},
    )


@require_POST
@login_required
def waiting_listattraction_join(request, pk=None):
    slot_id = request.POST.get("slot") or pk
    slot = get_object_or_404(
        VisitSlot.objects.select_related("attraction"),
        pk=slot_id
    )
    attraction_obj = slot.attraction
    if slot.remaining > 0:
        messages.error(
            request,
            f"{attraction_obj.name} still has availability for this slot. Please book directly."
        )
        return redirect("attraction", pk=attraction_obj.pk)

    existing = AttractionWaitlistEntry.objects.filter(
        user=request.user,
        slot=slot,
        cancelled=False,
    ).first()

    if existing:
        messages.info(
            request,
            f"You're already on the waiting list for {attraction_obj.name} on {slot.date} at {slot.time}."
        )
    else:
        AttractionWaitlistEntry.objects.create(
            user=request.user,
            attraction=attraction_obj,
            slot=slot,
        )
        messages.success(
            request,
            f"You joined the waiting list for {attraction_obj.name} on {slot.date} at {slot.time}."
        )

    return redirect("waiting_listattraction")

@require_POST
@login_required
def waiting_listattraction_leave(request, pk):
    slot = get_object_or_404(
        VisitSlot.objects.select_related("attraction"),
        pk=pk
    )
    attraction_obj = slot.attraction

    entry = AttractionWaitlistEntry.objects.filter(
        user=request.user,
        slot=slot,
        cancelled=False,
    ).first()

    if entry:
        entry.cancelled = True
        entry.save(update_fields=["cancelled"])
        messages.success(
            request,
            f"You left the waiting list for {attraction_obj.name} on {slot.date} at {slot.time}."
        )
    else:
        messages.info(
            request,
            f"You were not on the waiting list for {attraction_obj.name} on {slot.date} at {slot.time}."
        )

    return redirect(request.META.get("HTTP_REFERER", "waiting_listattraction"))


def reassign_cancelled_attraction_booking(slot, attraction_obj):
    """
    Reassign a newly freed attraction slot to the next eligible user
    on the waiting list for that exact slot.
    """
    waitlist_entries = (
        AttractionWaitlistEntry.objects
        .select_related("user")
        .filter(
            slot=slot,
            cancelled=False,
        )
        .order_by("created_at")
    )

    for entry in waitlist_entries:
        user = entry.user

        # check yearly booking limit
        remaining_allowance = calculate_remaining_allowance(user, "regular")
        if remaining_allowance <= 0:
            continue

        # prevent duplicate active booking for same slot
        already_booked = Booking.objects.filter(
            user=user,
            slot=slot,
            cancelled=False,
        ).exists()
        if already_booked:
            entry.cancelled = True
            entry.save(update_fields=["cancelled"])
            continue

        # ensure slot is still available
        slot.refresh_from_db()
        if slot.remaining < 1:
            return None

        new_booking = Booking.objects.create(
            user=user,
            attraction=attraction_obj,
            slot=slot,
            num_tickets=1,
            full_name=f"{user.first_name} {user.last_name}".strip(),
            email=user.email,
            agreed_terms=True,
        )

        VisitSlot.objects.filter(pk=slot.pk).update(
            remaining=F("remaining") - 1
        )

        entry.cancelled = True
        entry.save(update_fields=["cancelled"])

        send_attraction_booking_email_ticket_reallocaton(new_booking)
        return new_booking

    return None
# -----------------------------
# Suggestions (user) + export (staff)
# -----------------------------
@login_required
def create_attraction_suggestion(request):
    if request.method == "POST":
        form = AttractionSuggestionForm(request.POST)
        if form.is_valid():
            suggestion = form.save(commit=False)
            suggestion.submitted_by = request.user
            suggestion.save()
            messages.success(request, "Suggestion submitted successfully.")
            # prevents duplicate submit on refresh
            return redirect("create_attraction_suggestion")
        messages.error(request, "Please fix the errors below and try again.")
    else:
        form = AttractionSuggestionForm()

    return render(
        request,
        "fergusonbequest/attraction_suggestion_page.html",
        {"form": form},
    )


@staff_member_required
def export_suggestions_excel(request):
    qs = (
        AttractionSuggestion.objects
        .select_related("submitted_by")
        .order_by("-created_at")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Attraction Suggestions"

    headers = [
        "Name",
        "Location",
        "Website URL",
        "Description",
        "Why Recommended",
        "Status",
        "Submitted By",
        "Created At",
    ]
    ws.append(headers)

    for s in qs:
        ws.append([
            s.name,
            getattr(s, "location", "") or "",
            getattr(s, "website_url", "") or "",
            getattr(s, "description", "") or "",
            getattr(s, "why_recommended", "") or "",
            getattr(s, "status", "") or "",
            (s.submitted_by.username if s.submitted_by else ""),
            s.created_at.strftime("%Y-%m-%d %H:%M") if getattr(s, "created_at", None) else "",
        ])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    dv = DataValidation(
        type="list",
        formula1='"Pending,In progress,Implemented,Rejected"',
        allow_blank=True
    )
    ws.add_data_validation(dv)
    if ws.max_row >= 2:
        dv.add(f"F2:F{ws.max_row}")

    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 55)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="attraction_suggestions.xlsx"'
    wb.save(response)
    return response


# -----------------------------
# Discount Codes (staff page)
# -----------------------------
@staff_member_required
@require_http_methods(["GET", "POST"])
def discount_codes_page(request):
    """Admin page: create new discount codes and manage all existing ones."""
    now = timezone.now()

    if request.method == "POST":
        form = DiscountCodeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Discount code created successfully.")
            return redirect("discount_codes")
        messages.error(request, "Please fix the errors below.")
    else:
        form = DiscountCodeForm()

    all_discounts = DiscountCode.objects.all().order_by("-created_at")
    now = timezone.now()

    return render(
        request,
        "fergusonbequest/discount_codes.html",
        {
            "form": form,
            "all_discounts": all_discounts,
            "now": now,
        },
    )


@staff_member_required
def discount_code_edit(request, pk):
    """Admin: edit an existing discount code."""
    code = get_object_or_404(DiscountCode, pk=pk)
    if request.method == "POST":
        form = DiscountCodeForm(request.POST, instance=code)
        if form.is_valid():
            form.save()
            messages.success(request, f"Discount code '{code.code}' updated.")
            return redirect("discount_codes")
        messages.error(request, "Please fix the errors below.")
    else:
        form = DiscountCodeForm(instance=code)
    return render(request, "fergusonbequest/discount_code_edit.html", {"form": form, "code": code})


@staff_member_required
@require_POST
def discount_code_toggle(request, pk):
    """Admin: toggle is_active on a discount code."""
    code = get_object_or_404(DiscountCode, pk=pk)
    code.is_active = not code.is_active
    code.save(update_fields=["is_active"])
    state = "shown" if code.is_active else "hidden"
    messages.success(request, f"Discount code '{code.code}' is now {state}.")
    return redirect("discount_codes")


@staff_member_required
@require_POST
def discount_code_delete(request, pk):
    """Admin: permanently delete a discount code."""
    code = get_object_or_404(DiscountCode, pk=pk)
    title = code.code
    code.delete()
    messages.success(request, f"Discount code '{title}' deleted.")
    return redirect("discount_codes")


@login_required
def user_discount_codes(request):
    """Page showing currently active discount codes to logged-in users."""
    now = timezone.now()
    active_codes = DiscountCode.objects.filter(
        is_active=True,
        valid_from__lte=now,
        valid_until__gte=now,
    ).order_by("valid_until")
    return render(request, "fergusonbequest/user_discount_codes.html", {
        "active_codes": active_codes,
        "now": now,
    })

# -----------------------------
# Staff / Admin dashboard + management
# -----------------------------
@staff_member_required
def admin_dashboard(request):
    now = timezone.now()

    active_draws_count = sum(1 for d in TicketDraw.objects.all() if _call_is_open(d, now))
    open_venues_count = sum(1 for a in Attraction.objects.all() if _call_is_open(a, now))

    bookings_count = Booking.objects.filter(cancelled=False).count()
    # Bookings and ticket draw booking unsent and canceled
    bookings_needing_tickets_count = Booking.objects.filter(
        cancelled=False,
        ticket_sent=False
    ).aggregate(total=Sum("num_tickets"))["total"] or 0

    return render(
        request,
        "fergusonbequest/admin_dashboard.html",
        {
            "active_draws_count": active_draws_count,
            "open_venues_count": open_venues_count,
            "bookings_count": bookings_count,
            "bookings_needing_tickets_count": bookings_needing_tickets_count,
        },
    )


@staff_member_required
def admin_management(request):
    tab = request.GET.get("tab", "draws")
    q = (request.GET.get("q") or "").strip()

    sort_draws = request.GET.get("sort_draws", "close_date_desc")
    sort_attractions = request.GET.get("sort_attractions", "date_desc")

    draws_qs = TicketDraw.objects.annotate(entry_count=Count("ticketdrawbooking"))
    attractions_qs = Attraction.objects.all()

    if q:
        draws_qs = draws_qs.filter(name__icontains=q)
        attractions_qs = attractions_qs.filter(name__icontains=q)

    if sort_draws == "open_first":
        draws_qs = draws_qs.order_by("-booking_open", "-booking_close", "name")
    elif sort_draws == "close_date":
        draws_qs = draws_qs.order_by("booking_close", "name")
    elif sort_draws == "close_date_desc":
        draws_qs = draws_qs.order_by("-booking_close", "name")
    else:
        draws_qs = draws_qs.order_by("name")

    if sort_attractions == "date":
        attractions_qs = attractions_qs.order_by("booking_open", "name")
    elif sort_attractions == "date_desc":
        attractions_qs = attractions_qs.order_by("-booking_open", "name")
    else:
        attractions_qs = attractions_qs.order_by("name")

    now = timezone.now()
    for d in draws_qs:
        d.is_open_now = bool(_call_is_open(d, now))
        d.is_closed_now = not d.is_open_now

    return render(
        request,
        "fergusonbequest/admin_management.html",
        {
            "draws": draws_qs,
            "attractions": attractions_qs,
            "tab": tab,
            "q": q,
            "sort_draws": sort_draws,
            "sort_attractions": sort_attractions,
        },
    )


# Compatibility alias: urls.py may expect management_view
@staff_member_required
def management_view(request):
    return admin_management(request)


@staff_member_required
@require_POST
def run_draw(request, draw_id):
    draw = get_object_or_404(TicketDraw, pk=draw_id)

    now = timezone.now()
    if _call_is_open(draw, now):
        messages.error(request, "This draw is still open. You can only run it after it closes.")
        return redirect(f"{reverse('admin_management')}?tab=draws")

    if draw.winner_booking:
        messages.error(request, "This draw has already been run and cannot be run again.")
        return redirect(f"{reverse('admin_management')}?tab=draws")

    entries = list(
        TicketDrawBooking.objects.filter(ticket_draw=draw, cancelled=False).select_related("user")
    )

    if not entries:
        draw.winner_booking = None
        draw.winner_selected_at = None
        draw.save(update_fields=["winner_booking", "winner_selected_at"])
        messages.error(request, "No active entries for this draw.")
        return redirect(f"{reverse('admin_management')}?tab=draws")

    draw.winner_booking = random.choice(entries)
    draw.winner_selected_at = timezone.now()
    draw.save(update_fields=["winner_booking", "winner_selected_at"])

    winner = draw.winner_booking
    winner_name = winner.full_name or (winner.user.get_username() if winner.user else "Winner")

    messages.success(request, f"Winner selected: {winner_name}")
    return redirect(f"{reverse('admin_management')}?tab=draws")


@staff_member_required
@require_POST
def mng_delete_ticket_draw(request, draw_id):
    draw = get_object_or_404(TicketDraw, pk=draw_id)
    draw.delete()
    messages.success(request, "Draw deleted.")
    return redirect("/admin-dashboard/management/?tab=draws")


@staff_member_required
@require_POST
def mng_delete_draw(request, draw_id):
    # Backwards-compatible alias for older URL patterns/tests
    return mng_delete_ticket_draw(request, draw_id)


@staff_member_required
@require_POST
def mng_delete_attraction(request, attraction_id):
    attraction_obj = get_object_or_404(Attraction, pk=attraction_id)
    attraction_obj.delete()
    messages.success(request, "Attraction deleted.")
    return redirect("/admin-dashboard/management/?tab=attractions")


@staff_member_required
def create_attraction(request):
    if request.method == "POST":
        form = AttractionCreateForm(request.POST, request.FILES)
        if form.is_valid():
            attraction_obj = form.save()
            messages.success(request, f'Attraction "{attraction_obj.name}" created successfully!')
            return redirect(f"{reverse('admin_management')}?tab=attractions")
    else:
        form = AttractionCreateForm()

    return render(
        request,
        "fergusonbequest/create_attraction.html",
        {"form": form, "title": "Create New Attraction"},
    )


@staff_member_required
def edit_attraction(request, pk):
    attraction_obj = get_object_or_404(Attraction, pk=pk)

    if request.method == "POST":
        form = AttractionCreateForm(request.POST, request.FILES, instance=attraction_obj)
        if form.is_valid():
            form.save()
            messages.success(request, f'Attraction "{attraction_obj.name}" updated successfully!')
            return redirect("admin_dashboard")
    else:
        form = AttractionCreateForm(instance=attraction_obj)

    return render(
        request,
        "fergusonbequest/edit_attraction.html",
        {"form": form, "title": "Edit Attraction", "attraction": attraction_obj},
    )


@staff_member_required
def create_ticket_draw(request):
    if request.method == "POST":
        form = TicketDrawCreateForm(request.POST, request.FILES)
        if form.is_valid():
            draw = form.save()
            messages.success(request, f'Ticket Draw "{draw.name}" created successfully!')
            return redirect(f"{reverse('admin_management')}?tab=draws")
    else:
        form = TicketDrawCreateForm()

    return render(
        request,
        "fergusonbequest/create_ticket_draw.html",
        {"form": form, "title": "Create New Ticket Draw"},
    )


@staff_member_required
def edit_ticket_draw(request, pk):
    draw = get_object_or_404(TicketDraw, pk=pk)

    if request.method == "POST":
        form = TicketDrawCreateForm(request.POST, request.FILES, instance=draw)
        if form.is_valid():
            form.save()
            messages.success(request, f'Ticket Draw "{draw.name}" updated successfully!')
            return redirect("admin_dashboard")
    else:
        form = TicketDrawCreateForm(instance=draw)

    return render(
        request,
        "fergusonbequest/edit_ticket_draw.html",
        {"form": form, "title": "Edit Ticket Draw", "ticket_draw": draw},
    )

# -----------------------------
# Reports (admin)
# -----------------------------
@staff_member_required
def admin_reports(request):
    name = request.GET.get("name")
    surname = request.GET.get("surname")
    guid = request.GET.get("guid")
    email = request.GET.get("email")
    start = request.GET.get("start")
    end = request.GET.get("end")
    venue = request.GET.get("venue")
    status = request.GET.get("status")
    q = request.GET.get("q")
    booking_type = request.GET.get("booking_type", "all")
    venue_select = request.GET.get("venue_select")
    specific_date = request.GET.get("specific_date")
    specific_time = request.GET.get("specific_time")
    date_select = request.GET.get("date_select")
    time_select = request.GET.get("time_select")
    ticket_code = request.GET.get("ticket_code")
    min_tickets = request.GET.get("min_tickets")
    max_tickets = request.GET.get("max_tickets")

    sort = request.GET.get("sort", "newest")

    today = timezone.localdate()

    def get_ticket_reference(booking):
        if not booking:
            return ""
        return (
                (getattr(booking, "generic_booking_code", None) or "").strip()
                or f"REF-{booking.id}"
        )

    # exclude converted draw bookings so they don't appear twice
    converted_booking_ids = TicketDrawBooking.objects.filter(
        converted_booking__isnull=False
    ).values_list("converted_booking_id", flat=True)

    draw_qs = TicketDrawBooking.objects.select_related(
        "user", "ticket_draw", "slot", "converted_booking"
    )

    attraction_qs = Booking.objects.select_related(
        "user", "attraction", "slot"
    ).exclude(id__in=converted_booking_ids)

    def apply_filters(qs_in, is_draw=True):
        qs_out = qs_in

        if name:
            qs_out = qs_out.filter(user__first_name__icontains=name)
        if surname:
            qs_out = qs_out.filter(user__last_name__icontains=surname)
        if guid:
            qs_out = qs_out.filter(user__profile__staff_guid__icontains=guid)
        if email:
            qs_out = qs_out.filter(user__email__icontains=email)
        if start:
            qs_out = qs_out.filter(slot__date__gte=start)
        if end:
            qs_out = qs_out.filter(slot__date__lte=end)

        if ticket_code:
            if is_draw:
                qs_out = qs_out.filter(
                    Q(converted_booking__ticket_code__icontains=ticket_code)
                    | Q(converted_booking__generic_booking_code__icontains=ticket_code)
                )
            else:
                qs_out = qs_out.filter(
                    Q(ticket_code__icontains=ticket_code)
                    | Q(generic_booking_code__icontains=ticket_code)
                )

        if min_tickets and min_tickets.isdigit():
            qs_out = qs_out.filter(num_tickets__gte=int(min_tickets))
        if max_tickets and max_tickets.isdigit():
            qs_out = qs_out.filter(num_tickets__lte=int(max_tickets))

        venue_value = venue if venue else venue_select
        if venue_value:
            if is_draw:
                qs_out = qs_out.filter(ticket_draw__name__icontains=venue_value)
            else:
                qs_out = qs_out.filter(attraction__name__icontains=venue_value)

        date_value = specific_date if specific_date else date_select
        if date_value:
            qs_out = qs_out.filter(slot__date=date_value)

        time_value = specific_time if specific_time else time_select
        if time_value:
            qs_out = qs_out.filter(slot__time=time_value)

        if status == "active":
            qs_out = qs_out.filter(cancelled=False, slot__date__gte=today)
        elif status == "cancelled":
            qs_out = qs_out.filter(cancelled=True)
        elif status == "completed":
            qs_out = qs_out.filter(cancelled=False, slot__date__lt=today)

        if q:
            common_filters = (
                Q(user__first_name__icontains=q)
                | Q(user__last_name__icontains=q)
                | Q(user__username__icontains=q)
                | Q(user__email__icontains=q)
                | Q(slot__date__icontains=q)
                | Q(slot__time__icontains=q)
                | Q(num_tickets__icontains=q)
            )

            if is_draw:
                common_filters |= (
                    Q(converted_booking__ticket_code__icontains=q)
                    | Q(converted_booking__generic_booking_code__icontains=q)
                )
                qs_out = qs_out.filter(common_filters | Q(ticket_draw__name__icontains=q))
            else:
                common_filters |= (
                    Q(ticket_code__icontains=q)
                    | Q(generic_booking_code__icontains=q)
                )
                qs_out = qs_out.filter(common_filters | Q(attraction__name__icontains=q))

        return qs_out

    combined = []

    if booking_type in ["all", "draw"]:
        for b in apply_filters(draw_qs, True):
            if b.cancelled:
                status_text = "Cancelled"
            elif b.slot.date < today:
                status_text = "Completed"
            else:
                status_text = "Active"

            combined.append(
                {
                    "type": "Draw",
                    "created": b.created_at,
                    "name": b.ticket_draw.name,
                    "first_name": b.user.first_name if b.user else "",
                    "last_name": b.user.last_name if b.user else "",
                    "guid": getattr(b.user.profile, "staff_guid", "") if b.user and hasattr(b.user, "profile") else "",
                    "email": b.user.email if b.user else (b.email or ""),
                    "date": b.slot.date,
                    "time": b.slot.time,
                    "ticket_code": get_ticket_reference(b.converted_booking),
                    "num_tickets": b.num_tickets,
                    "status_text": status_text,
                }
            )

    if booking_type in ["all", "attraction"]:
        for b in apply_filters(attraction_qs, False):
            if b.cancelled:
                status_text = "Cancelled"
            elif b.slot.date < today:
                status_text = "Completed"
            else:
                status_text = "Active"

            combined.append(
                {
                    "type": "Attraction",
                    "created": b.created_at,
                    "name": b.attraction.name,
                    "first_name": b.user.first_name if b.user else "",
                    "last_name": b.user.last_name if b.user else "",
                    "guid": getattr(b.user.profile, "staff_guid", "") if b.user and hasattr(b.user, "profile") else "",
                    "email": b.user.email if b.user else (b.email or ""),
                    "date": b.slot.date,
                    "time": b.slot.time,
                    "ticket_code": get_ticket_reference(b),
                    "num_tickets": b.num_tickets,
                    "status_text": status_text,
                }
            )

    combined.sort(key=itemgetter("created"), reverse=(sort == "newest"))

    # Pagination
    page = request.GET.get("page", 1)
    per_page = 20
    total_bookings = len(combined)
    total_pages = (total_bookings + per_page - 1) // per_page

    try:
        page = int(page)
        if page < 1:
            page = 1
        elif total_pages > 0 and page > total_pages:
            page = total_pages
    except (ValueError, TypeError):
        page = 1

    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    paginated_bookings = combined[start_idx:end_idx]

    export_type = request.GET.get("export")
    if export_type:
        headers = [
            "Type",
            "Attraction/Draw",
            "Forename",
            "Surname",
            "GUID",
            "Email",
            "Date",
            "Time",
            "Ticket Reference",
            "Tickets",
            "Status",
        ]

        if export_type == "csv":
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = 'attachment; filename="reports.csv"'
            writer = csv.writer(response)
            writer.writerow(headers)

            for b in combined:
                writer.writerow(
                    [
                        b["type"],
                        b["name"],
                        b["first_name"],
                        b["last_name"],
                        b["guid"],
                        b["email"],
                        b["date"].strftime("%d/%m/%Y"),
                        b["time"].strftime("%H:%M") if b["time"] else "",
                        b["ticket_code"],
                        b["num_tickets"],
                        b["status_text"],
                    ]
                )
            return response

        if export_type == "excel":
            wb = Workbook()
            ws = wb.active
            ws.title = "Reports"

            ws.append(headers)

            for b in combined:
                ws.append(
                    [
                        b["type"],
                        b["name"],
                        b["first_name"],
                        b["last_name"],
                        b["guid"],
                        b["email"],
                        b["date"].strftime("%d/%m/%Y"),
                        b["time"].strftime("%H:%M") if b["time"] else "",
                        b["ticket_code"],
                        b["num_tickets"],
                        b["status_text"],
                    ]
                )

            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="reports.xlsx"'
            wb.save(response)
            return response

    def calculate_statistics(bookings_list):
        total = len(bookings_list)
        attraction_count = sum(1 for b in bookings_list if b["type"] == "Attraction")
        draw_count = total - attraction_count

        active_count = sum(1 for b in bookings_list if b["status_text"] == "Active")
        completed_count = sum(1 for b in bookings_list if b["status_text"] == "Completed")
        cancelled_count = sum(1 for b in bookings_list if b["status_text"] == "Cancelled")

        total_tickets = sum(b.get("num_tickets", 0) for b in bookings_list)

        date_range = None
        if bookings_list:
            dates = [b["date"] for b in bookings_list]
            date_range = {"start": min(dates), "end": max(dates)}

        popularity = {}
        for b in bookings_list:
            popularity[b["name"]] = popularity.get(b["name"], 0) + 1

        most_popular = None
        if popularity:
            n, c = max(popularity.items(), key=lambda x: x[1])
            most_popular = {"name": n, "count": c}

        unique_users = len(set(b["email"] for b in bookings_list if b["email"]))
        avg_per_user = total / unique_users if unique_users > 0 else 0

        return {
            "total_bookings": total,
            "total_tickets": total_tickets,
            "attraction_count": attraction_count,
            "draw_count": draw_count,
            "active_count": active_count,
            "completed_count": completed_count,
            "cancelled_count": cancelled_count,
            "date_range": date_range,
            "most_popular": most_popular,
            "unique_users": unique_users,
            "avg_per_user": avg_per_user,
        }

    statistics = calculate_statistics(combined)

    venue_list = sorted({b["name"] for b in combined})
    date_list = sorted({b["date"] for b in combined}, reverse=True)
    time_list = sorted({b["time"] for b in combined if b["time"]})

    return render(
        request,
        "fergusonbequest/admin_reports.html",
        {
            "bookings": paginated_bookings,
            "selected_booking_type": booking_type,
            "selected_status": status,
            "statistics": statistics,
            "current_page": page,
            "total_pages": total_pages,
            "total_bookings": total_bookings,
            "has_previous": page > 1,
            "has_next": page < total_pages,
            "previous_page": page - 1,
            "next_page": page + 1,
            "page_range": range(1, total_pages + 1),
            "start_index": (start_idx + 1) if total_bookings else 0,
            "end_index": min(end_idx, total_bookings),
            "venue_list": venue_list,
            "date_list": date_list,
            "time_list": time_list,
        },
    )


# -----------------------------
# Emails (admin)
# -----------------------------
@staff_member_required
def admin_email(request):
    email_types = EmailTemplate.TYPE_CHOICES

    selected_type = request.GET.get("email_type")
    template_id = request.GET.get("template_id")

    templates = EmailTemplate.objects.filter(type=selected_type) if selected_type else []

    selected_template = None
    if template_id:
        selected_template = get_object_or_404(EmailTemplate, id=template_id)

    feedback_settings = FeedbackEmailTemplate.get_template() if selected_type == "feedback" else None

    # POST actions
    if request.method == "POST" and selected_template:

        subject = request.POST.get("subject")
        body = request.POST.get("body")

        # SAVE
        if "save" in request.POST:
            selected_template.subject = subject
            selected_template.body = body
            selected_template.save()
            
            if selected_template.type == "feedback":
                feedback_singleton = FeedbackEmailTemplate.get_template()
                feedback_singleton.subject = subject
                feedback_singleton.body = body
                feedback_singleton.enabled = request.POST.get("feedback_enabled") == "on"
                feedback_singleton.feedback_url = request.POST.get("feedback_url", "").strip()
                feedback_singleton.save(update_fields=["subject", "body", "enabled", "feedback_url"])
            
            messages.success(request, "Template saved")

        # SEND TEST
        elif "send" in request.POST:
            context = get_email_context(user=request.user)
            if selected_template.type == "feedback":
                feedback_singleton = FeedbackEmailTemplate.get_template()
                context["feedback_url"] = feedback_singleton.feedback_url
            send_template_email(
                selected_template.type,
                request.user.email,  # send to yourself
                context,
            )
            messages.success(request, "Test email sent")

        elif "send_feedback_now" in request.POST and selected_template.type == "feedback":
            from .scheduler import send_scheduled_feedback_emails
            send_scheduled_feedback_emails()
            messages.success(request, "Feedback emails job ran successfully")

        elif "send_announcement_all" in request.POST:
            selected_template.subject = subject
            selected_template.body = body
            selected_template.save()
            all_users = User.objects.all()
            count = 0
            for user in all_users:
                try:
                    send_announcement_email(user)
                    count += 1
                except Exception as e:
                    pass
            messages.success(request, f"Announcement sent to all {count} users.")

        elif "send_custom_selected" in request.POST:
            selected_template.subject = subject
            selected_template.body = body
            selected_template.save()
            raw_ids = request.POST.get("custom_recipient_ids", "")
            ids = [int(i) for i in raw_ids.split(",") if i.strip().isdigit()]
            users = User.objects.filter(id__in=ids)
            count = 0
            for user in users:
                try:
                    send_custom_email(user)
                    count += 1
                except Exception:
                    pass
            messages.success(request, f"Custom email sent to {count} selected user(s).")


        elif "set_default" in request.POST:
            EmailTemplate.objects.filter(type=selected_template.type).update(is_default=False)
            selected_template.is_default = True
            selected_template.save()
            messages.success(request,
                             f"'{selected_template.name}' is now the default template for {selected_template.get_type_display()}")

        # DELETE
        elif "delete" in request.POST:
            was_default = selected_template.is_default
            template_name = selected_template.name
            template_type = selected_template.type

            selected_template.delete()

            # If we deleted the default, set another template as default if available
            if was_default:
                next_template = EmailTemplate.objects.filter(type=template_type).first()
                if next_template:
                    next_template.is_default = True
                    next_template.save()
                    messages.success(request, f"Template deleted. '{next_template.name}' is now the default.")
                else:
                    messages.success(request, "Template deleted. No templates remain for this type.")
            else:
                messages.success(request, "Template deleted")

            return redirect(request.path + f"?email_type={selected_type}")

    # CREATE NEW
    if request.method == "POST" and "create" in request.POST:
        is_first = not EmailTemplate.objects.filter(type=selected_type).exists()

        new_template = EmailTemplate.objects.create(
            name="New Template",
            type=selected_type,
            subject="Subject here",
            body="",
            is_default=is_first  # First template becomes default automatically
        )

        messages.success(request, "Template created" + (" and set as default" if is_first else ""))
        return redirect(request.path + f"?email_type={selected_type}&template_id={new_template.id}")

    context = {
        "email_types": email_types,
        "templates": templates,
        "selected_template": selected_template,
        "selected_type": selected_type,
        "feedback_settings": feedback_settings,
        "all_users": User.objects.all().order_by("last_name", "first_name"),
    }

    return render(request, "fergusonbequest/admin_email.html", context)


def send_template_email(template_type, recipient, context_dict, attachments=None):
    template = (
        EmailTemplate.objects.filter(type=template_type, is_default=True).first()
        or EmailTemplate.objects.filter(type=template_type).first()
    )

    if not template:
        return
        
    subject = Template(template.subject).render(Context(context_dict))
    body_content = Template(template.body).render(Context(context_dict))
    
    html_body = render_to_string(
        "fergusonbequest/base_email.html",
        {"body": body_content}
    )
    
    email = EmailMultiAlternatives(
        subject,
        body_content,
        settings.DEFAULT_FROM_EMAIL,
        [recipient]
    )
    
    email.attach_alternative(html_body, "text/html")
    
    # Add file attachments if provided
    if attachments:
        for attachment in attachments:
            try:
                # Handle BookingTicket objects
                if isinstance(attachment, BookingTicket) and attachment.file:
                    with attachment.file.open('rb') as f:
                        file_content = f.read()
                        filename = attachment.file.name.split('/')[-1]
                        
                        # Determine mime type based on file extension
                        if filename.lower().endswith('.pdf'):
                            mime_type = 'application/pdf'
                        elif filename.lower().endswith('.png'):
                            mime_type = 'image/png'
                        elif filename.lower().endswith(('.jpg', '.jpeg')):
                            mime_type = 'image/jpeg'
                        elif filename.lower().endswith('.gif'):
                            mime_type = 'image/gif'
                        else:
                            mime_type = 'application/octet-stream'
                            
                        email.attach(filename, file_content, mime_type)
                        
            except Exception as e:
                continue
    
    email.send(fail_silently=False)


# Confirmation
def send_attraction_booking_email_confirmation(booking):
    context = get_email_context(booking=booking)
    send_template_email(
        "attraction_confirmation",
        booking.user.email,
        context,
    )


def send_draw_booking_email_confirmation(draw_booking):
    context = get_email_context(draw_booking=draw_booking)
    send_template_email(
        "draw_confirmation",
        draw_booking.user.email,
        context,
    )


# Cancellation
def send_attraction_booking_email_cancellation(booking):
    context = get_email_context(booking=booking)
    send_template_email(
        "attraction_cancellation",
        booking.user.email,
        context,
    )


def send_draw_booking_email_cancellation(draw_booking):
    context = get_email_context(draw_booking=draw_booking)
    send_template_email(
        "draw_cancellation",
        draw_booking.user.email,
        context,
    )


# Ticket Distribution (send after deadline date)
def send_attraction_booking_email_ticket_distribution(booking):
    # Collect ticket files as attachments
    attachments = []
    
    # Get all ticket files associated with this booking
    if booking.tickets.exists():
        attachments = list(booking.tickets.all())
    
    context = get_email_context(booking=booking)
    send_template_email(
        "attraction_distribution",
        booking.user.email,
        context,
        attachments=attachments  # Pass attachments
    )


# sent after accepting the ticket
def send_draw_booking_email_ticket_distribution(draw_booking):
    # Collect ticket files as attachments
    attachments = []
    
    # If accepted and converted, get tickets from converted booking
    if draw_booking.is_accepted and draw_booking.converted_booking:
        converted = draw_booking.converted_booking
        if converted.tickets.exists():
            attachments = list(converted.tickets.all())
    
    context = get_email_context(draw_booking=draw_booking)
    send_template_email(
        "draw_distribution",
        draw_booking.user.email,
        context,
        attachments=attachments  # Pass attachments
    )

# Ticket Draw Winner (Accept or Decline, 3 day deadline)
def send_draw_booking_email_winner(draw_booking):
    deadline = timezone.now() + datetime.timedelta(hours=72)

    context = get_email_context(draw_booking=draw_booking, winner_deadline=deadline)
    send_template_email(
        "draw_winner",
        draw_booking.user.email,
        context,
    )


# Attraction List Reallocation
def send_attraction_booking_email_ticket_reallocaton(booking):
    context = get_email_context(booking=booking)
    send_template_email(
        "attraction_reallocation",
        booking.user.email,
        context,
    )


# Ticket Draw Redraw Winner
def send_draw_booking_email_redraw_winner(draw_booking):
    deadline = timezone.now() + datetime.timedelta(hours=72)

    context = get_email_context(draw_booking=draw_booking, winner_deadline=deadline)
    send_template_email(
        "draw_reallocation",
        draw_booking.user.email,
        context,
    )


# Reminder (1 day before)
def send_attraction_booking_email_reminder(booking):
    context = get_email_context(booking=booking)
    send_template_email(
        "attraction_reminder",
        booking.user.email,
        context,
    )


def send_draw_booking_email_reminder(draw_booking):
    context = get_email_context(draw_booking=draw_booking)
    send_template_email(
        "draw_reminder",
        draw_booking.user.email,
        context,
    )


def send_announcement_email(user):
    context = get_email_context(user=user)
    send_template_email(
        "announcement",
        user.email,
        context,
    )


def send_custom_email(user):
    context = get_email_context(user=user)
    send_template_email(
        "custom",
        user.email,
        context,
    )


def send_feedback_email_request(booking, feedback_url):
    recipient = booking.user.email if booking.user and booking.user.email else booking.email
    context = get_email_context(booking=booking, feedback_url=feedback_url)
    send_template_email(
        "feedback",
        recipient,
        context,
    )
    




def get_email_context(booking=None, draw_booking=None, user=None, **kwargs):
    try:
        from django.contrib.sites.models import Site
        current_site = Site.objects.get_current()

        if '127.0.0.1' in current_site.domain or 'localhost' in current_site.domain:
            domain = f"http://{current_site.domain}"
        else:
            domain = f"https://{current_site.domain}"

    except:
        domain = "http://127.0.0.1:8000"

    context = {
        "current_date": timezone.now().strftime("%d/%m/%Y"),
        "current_time": timezone.now().strftime("%H:%M"),
        "site_url": "https://www.gla.ac.uk/myglasgow/courtoffice/fergusonbequest/",
        "homepage_url": domain,
        "contact_email": "fergusonbequest@glasgow.ac.uk",
    }

    if user:
        context.update({
            "user_id": user.id,
            "username": user.username,
            "email": user.email,
            "first_name": user.first_name,
            "last_name": user.last_name,
            "full_name": f"{user.first_name} {user.last_name}".strip(),
        })

    elif booking and booking.user:
        context.update({
            "user_id": booking.user.id,
            "username": booking.user.username,
            "email": booking.user.email,
            "first_name": booking.user.first_name,
            "last_name": booking.user.last_name,
            "full_name": booking.full_name or f"{booking.user.first_name} {booking.user.last_name}".strip(),
        })

    elif draw_booking and draw_booking.user:
        context.update({
            "user_id": draw_booking.user.id,
            "username": draw_booking.user.username,
            "email": draw_booking.user.email,
            "first_name": draw_booking.user.first_name,
            "last_name": draw_booking.user.last_name,
            "full_name": draw_booking.full_name or f"{draw_booking.user.first_name} {draw_booking.user.last_name}".strip(),
        })

    if booking:

        slot = booking.slot if hasattr(booking, 'slot') else None

        visit_time = ""
        visit_date = ""
        visit_day = ""
        visit_datetime = ""

        if slot:
            if hasattr(slot, 'time') and slot.time:
                visit_time = slot.time.strftime("%H:%M")
            if hasattr(slot, 'date') and slot.date:
                visit_date = slot.date.strftime("%d/%m/%Y")
                visit_day = slot.date.strftime("%A")
            if visit_date and visit_time:
                visit_datetime = f"{visit_date} at {visit_time}"

        ticket_files = booking.tickets.all().order_by('sort_order')
        
        ticket_codes_list = []  # For "codes" type
        pdf_tickets = []        # For "pdf_template" and "pdf_template_random"
        qr_tickets = []         # For "qr_individual"
        booking_code = ""       # For "booking_code" type
        entry_instructions = "" # For "instructions" type
        box_office_notes = ""   # For "box_office"
        
        # Handle each ticket type specifically
        if booking.ticket_type == "codes" and booking.ticket_code:
            ticket_codes_list = booking.ticket_code.split('\n')
        
        elif booking.ticket_type == "qr_individual":
            qr_tickets = [
                {
                    'file_url': f"{domain}{t.file.url}" if t.file and t.file.url.startswith('/') else (t.file.url if t.file else ''),
                    'qr_value': t.qr_value,
                    'filename': t.file.name.split('/')[-1] if t.file else '',
                }
                for t in ticket_files if t.file
            ]
        
        elif booking.ticket_type in ["pdf_template", "pdf_template_random"]:
            pdf_tickets = [
                {
                    'file_url': f"{domain}{t.file.url}" if t.file and t.file.url.startswith('/') else (t.file.url if t.file else ''),
                    'ticket_code': t.ticket_code,
                    'filename': t.file.name.split('/')[-1] if t.file else '',
                }
                for t in ticket_files if t.file
            ]
        
        elif booking.ticket_type == "booking_code" and booking.generic_booking_code:
            return HttpResponse(
                f"Booking Code: {booking.generic_booking_code}",
                content_type="text/plain",
            )
        
        elif booking.ticket_type == "instructions" and booking.ticket_instructions:
            entry_instructions = booking.ticket_instructions
        
        elif booking.ticket_type == "box_office":
            box_office_notes = booking.box_office_notes or ""

        total_tickets_count = (
            len(ticket_codes_list) + 
            len(pdf_tickets) + 
            len(qr_tickets) + 
            (1 if booking_code else 0) + 
            (1 if entry_instructions else 0) + 
            (1 if booking.ticket_type == "box_office" else 0)
        )

        # Determine which ticket type is active and set all_tickets to that
        all_tickets = []
        if ticket_codes_list:
            all_tickets = [{"type": "code", "value": code} for code in ticket_codes_list]
        elif pdf_tickets:
            all_tickets = pdf_tickets
        elif qr_tickets:
            all_tickets = qr_tickets
        elif booking_code:
            all_tickets = [{"type": "booking_code", "code": booking_code}]
        elif entry_instructions:
            all_tickets = [{"type": "instructions", "text": entry_instructions}]
        elif booking.ticket_type == "box_office":
            all_tickets = [{"type": "box_office", "notes": box_office_notes}]

        # Format PDF tickets as HTML string
        pdf_tickets_html = ""
        if pdf_tickets:
            pdf_tickets_html = "<div class='pdf-tickets'>"
            for i, pdf in enumerate(pdf_tickets, 1):
                pdf_tickets_html += f"""
                <div class='ticket' style='margin-bottom: 15px; padding: 10px; border: 1px solid #ddd; border-radius: 4px;'>
                    <h4 style='margin-top: 0; color: #002663;'>PDF Ticket {i}</h4>
                    <p><a href='{pdf['file_url']}' style='background-color: #002663; color: white; padding: 8px 15px; text-decoration: none; border-radius: 4px; display: inline-block;'>Download {pdf['filename']}</a></p>
                    {f"<p><strong>Code:</strong> {pdf['ticket_code']}</p>" if pdf['ticket_code'] else ""}
                </div>
                """
            pdf_tickets_html += "</div>"
        
        # Format ticket codes as HTML string
        ticket_codes_html = ""
        if ticket_codes_list:
            ticket_codes_html = "<div class='code-tickets'><ul style='list-style-type: none; padding: 0;'>"
            for code in ticket_codes_list:
                ticket_codes_html += f"<li style='margin-bottom: 5px; padding: 8px; background-color: #f5f5f5; border-radius: 4px;'><strong>Code:</strong> {code}</li>"
            ticket_codes_html += "</ul></div>"
        
        # Format QR tickets as HTML string
        qr_tickets_html = ""
        if qr_tickets:
            qr_tickets_html = "<div class='qr-tickets'>"
            for i, qr in enumerate(qr_tickets, 1):
                qr_tickets_html += f"""
                <div class='ticket' style='margin-bottom: 15px; padding: 10px; border: 1px solid #ddd; border-radius: 4px;'>
                    <h4 style='margin-top: 0; color: #002663;'>QR Ticket {i}</h4>
                    {f"<p><strong>QR Value:</strong> {qr['qr_value']}</p>" if qr['qr_value'] else ""}
                    <p><img src='{qr['file_url']}' alt='QR Code' style='max-width: 200px; border: 1px solid #ddd; padding: 5px;'></p>
                    <p><small>Filename: {qr['filename']}</small></p>
                </div>
                """
            qr_tickets_html += "</div>"
        
        # Format all_tickets as HTML string
        all_tickets_html = ""
        if all_tickets:
            all_tickets_html = "<div class='all-tickets'>"
            for i, ticket in enumerate(all_tickets, 1):
                # Check if it's a code ticket
                if ticket.get('type') == 'code' or (isinstance(ticket, dict) and 'value' in ticket and not ticket.get('file_url')):
                    value = ticket.get('value', '')
                    all_tickets_html += f"""
                    <div class='ticket' style='margin-bottom: 15px; padding: 10px; border: 1px solid #ddd; border-radius: 4px;'>
                        <h4 style='margin-top: 0; color: #002663;'>Code Ticket {i}</h4>
                        <p><strong>Code:</strong> {value}</p>
                    </div>
                    """
                
                # Check if it's a QR ticket
                elif ticket.get('type') == 'qr' or ticket.get('qr_value') is not None:
                    qr_value = ticket.get('qr_value', '')
                    file_url = ticket.get('file_url', '')
                    filename = ticket.get('filename', f'qr-code-{i}.png')
                    
                    all_tickets_html += f"""
                    <div class='ticket' style='margin-bottom: 15px; padding: 10px; border: 1px solid #ddd; border-radius: 4px;'>
                        <h4 style='margin-top: 0; color: #002663;'>QR Ticket {i}</h4>
                        {f"<p><strong>QR Value:</strong> {qr_value}</p>" if qr_value else ""}
                        {f"<p><img src='{file_url}' alt='QR Code' style='max-width: 200px; border: 1px solid #ddd; padding: 5px;'></p>" if file_url else ""}
                        <p><small>Filename: {filename}</small></p>
                    </div>
                    """
                
                # Check if it's a PDF ticket
                elif ticket.get('file_url') and (ticket.get('filename', '').lower().endswith('.pdf') or ticket.get('type') == 'pdf'):
                    file_url = ticket.get('file_url', '')
                    filename = ticket.get('filename', f'pdf-ticket-{i}.pdf')
                    ticket_code = ticket.get('ticket_code', '')
                    
                    all_tickets_html += f"""
                    <div class='ticket' style='margin-bottom: 15px; padding: 10px; border: 1px solid #ddd; border-radius: 4px;'>
                        <h4 style='margin-top: 0; color: #002663;'>PDF Ticket {i}</h4>
                        <p><a href='{file_url}' style='background-color: #002663; color: white; padding: 8px 15px; text-decoration: none; border-radius: 4px; display: inline-block;'>Download {filename}</a></p>
                        {f"<p><strong>Code:</strong> {ticket_code}</p>" if ticket_code else ""}
                    </div>
                    """
                
                # Check if it's a file ticket
                elif ticket.get('file_url'):
                    file_url = ticket.get('file_url', '')
                    filename = ticket.get('filename', f'file-{i}')
                    ticket_code = ticket.get('ticket_code', '')
                    
                    all_tickets_html += f"""
                    <div class='ticket' style='margin-bottom: 15px; padding: 10px; border: 1px solid #ddd; border-radius: 4px;'>
                        <h4 style='margin-top: 0; color: #002663;'>Ticket {i}</h4>
                        <p><a href='{file_url}' style='background-color: #002663; color: white; padding: 8px 15px; text-decoration: none; border-radius: 4px; display: inline-block;'>Download {filename}</a></p>
                        {f"<p><strong>Code:</strong> {ticket_code}</p>" if ticket_code else ""}
                    </div>
                    """
                
                # Check if it's a booking code
                elif ticket.get('type') == 'booking_code' or ticket.get('code'):
                    code = ticket.get('code', ticket.get('value', ''))
                    all_tickets_html += f"""
                    <div class='ticket' style='margin-bottom: 15px; padding: 15px; background-color: #f5f5f5; border-radius: 4px; text-align: center;'>
                        <h4 style='margin-top: 0; color: #002663;'>Booking Code</h4>
                        <p style='font-size: 18px; font-weight: bold; color: #002663;'>{code}</p>
                    </div>
                    """
                
                # Check if it's instructions
                elif ticket.get('type') == 'instructions' or ticket.get('text'):
                    text = ticket.get('text', ticket.get('value', ''))
                    all_tickets_html += f"""
                    <div class='ticket' style='margin-bottom: 15px; padding: 15px; background-color: #f5f5f5; border-radius: 4px;'>
                        <h4 style='margin-top: 0; color: #002663;'>Entry Instructions</h4>
                        <p>{text}</p>
                    </div>
                    """
                
                # Check if it's box office
                elif ticket.get('type') == 'box_office' or ticket.get('notes'):
                    notes = ticket.get('notes', ticket.get('value', ''))
                    all_tickets_html += f"""
                    <div class='ticket' style='margin-bottom: 15px; padding: 15px; background-color: #f5f5f5; border-radius: 4px;'>
                        <h4 style='margin-top: 0; color: #002663;'>Box Office Collection</h4>
                        <p>{notes if notes else 'Please collect your tickets at the box office.'}</p>
                    </div>
                    """
                
                # Fallback for any other type
                else:
                    all_tickets_html += f"<div style='margin-bottom: 10px; padding: 8px; background-color: #f5f5f5; border-radius: 4px;'><pre>{ticket}</pre></div>"
                    
            all_tickets_html += "</div>"
        
        # Format booking code as HTML
        booking_code_html = ""
        if booking_code:
            booking_code_html = f"<div style='margin-bottom: 15px; padding: 15px; background-color: #f5f5f5; border-radius: 4px; text-align: center;'><h4 style='margin-top: 0; color: #002663;'>Booking Code</h4><p style='font-size: 18px; font-weight: bold; color: #002663;'>{booking_code}</p></div>"
        
        # Format entry instructions as HTML
        entry_instructions_html = ""
        if entry_instructions:
            entry_instructions_html = f"<div style='margin-bottom: 15px; padding: 15px; background-color: #f5f5f5; border-radius: 4px;'><h4 style='margin-top: 0; color: #002663;'>Entry Instructions</h4><p>{entry_instructions}</p></div>"
        
        # Format box office notes as HTML
        box_office_html = ""
        if booking.ticket_type == "box_office":
            box_office_html = f"<div style='margin-bottom: 15px; padding: 15px; background-color: #f5f5f5; border-radius: 4px;'><h4 style='margin-top: 0; color: #002663;'>Box Office Collection</h4><p>{box_office_notes if box_office_notes else 'Please collect your tickets at the box office.'}</p></div>"

        from django.utils.safestring import mark_safe
        context.update({

            "booking_type": "attraction",
            "booking_id": booking.id,
            "booking_created_date": booking.created_at.strftime("%d/%m/%Y %H:%M") if booking.created_at else "",
            "booking_status": "Cancelled" if booking.cancelled else "Confirmed",
            "booking_num_tickets": booking.num_tickets,

            "attraction_id": booking.attraction.id if booking.attraction else "",
            "attraction_name": booking.attraction.name if booking.attraction else "",
            "attraction_location": booking.attraction.location if booking.attraction else "",
            "attraction_description": booking.attraction.description if booking.attraction else "",

            "visit_date": visit_date,
            "visit_day": visit_day,
            "visit_time": visit_time,
            "visit_datetime": visit_datetime,

            "cancel_link": f"{domain}/booking/{booking.id}/cancel/",
            "view_booking_link": f"{domain}/booking/history/#booking-{booking.id}",

            #ticket (metadata)
            "ticket_type": booking.ticket_type,
            "ticket_sent_date": booking.ticket_sent_at.strftime("%d/%m/%Y %H:%M") if booking.ticket_sent_at else "",

            #all tickets
            "all_tickets": mark_safe(all_tickets_html),
            "all_tickets_count": total_tickets_count,

            #ticket codes
            "ticket_codes": mark_safe(ticket_codes_html),
            "ticket_codes_count": len(ticket_codes_list),

            #pdf tickets (file + code)
            "pdf_tickets": mark_safe(pdf_tickets_html),
            "pdf_tickets_count": len(pdf_tickets),

            #qr ticket (file + qr value)
            "qr_tickets": mark_safe(qr_tickets_html),
            "qr_tickets_count": len(qr_tickets),

            #booking code
            "booking_code": mark_safe(booking_code_html),
        
            #entry instruction
            "entry_instructions": mark_safe(entry_instructions_html),

            #box office collection
            "box_office_instructions": mark_safe(box_office_html),
            
        })

    if draw_booking:

        slot = draw_booking.slot if hasattr(draw_booking, 'slot') else None

        visit_time = ""
        visit_date = ""
        visit_day = ""
        visit_datetime = ""

        if slot:
            if hasattr(slot, 'time') and slot.time:
                visit_time = slot.time.strftime("%H:%M")
            if hasattr(slot, 'date') and slot.date:
                visit_date = slot.date.strftime("%d/%m/%Y")
                visit_day = slot.date.strftime("%A")
            if visit_date and visit_time:
                visit_datetime = f"{visit_date} at {visit_time}"

        context.update({

            "booking_type": "draw",
            "booking_id": draw_booking.id,
            "booking_created_date": draw_booking.created_at.strftime(
                "%d/%m/%Y %H:%M") if draw_booking.created_at else "",
            "booking_status": "Cancelled" if draw_booking.cancelled else "Entered",
            "booking_num_tickets": draw_booking.num_tickets,

            "draw_id": draw_booking.ticket_draw.id if draw_booking.ticket_draw else "",
            "draw_name": draw_booking.ticket_draw.name if draw_booking.ticket_draw else "",
            "draw_location": draw_booking.ticket_draw.location if draw_booking.ticket_draw else "",
            "draw_description": draw_booking.ticket_draw.description if draw_booking.ticket_draw else "",

            "visit_date": visit_date,
            "visit_day": visit_day,
            "visit_time": visit_time,
            "visit_datetime": visit_datetime,

            "cancel_link": f"{domain}/ticket-draw/{draw_booking.id}/cancel/",
            "view_booking_link": f"{domain}/booking/history/#booking-{draw_booking.id}",
            "accept_link": f"{domain}/draw/accept/{draw_booking.id}",
            "reject_link": f"{domain}/draw/decline/{draw_booking.id}",
        })

        # if accepted and converted
        if draw_booking.is_accepted and draw_booking.converted_booking:
            converted = draw_booking.converted_booking
            
            converted_ticket_files = converted.tickets.all().order_by('sort_order')
            
            # Initialize all ticket type variables for converted booking
            converted_ticket_codes_list = []
            converted_pdf_tickets = []
            converted_qr_tickets = []
            converted_booking_code = ""
            converted_entry_instructions = ""
            converted_box_office_notes = ""
            
            # Handle each ticket type specifically
            if converted.ticket_type == "codes" and converted.ticket_code:
                converted_ticket_codes_list = converted.ticket_code.split('\n')
            
            elif converted.ticket_type == "qr_individual":
                converted_qr_tickets = [
                    {
                        'file_url': f"{domain}{t.file.url}" if t.file and t.file.url.startswith('/') else (t.file.url if t.file else ''),
                        'qr_value': t.qr_value,
                        'filename': t.file.name.split('/')[-1] if t.file else '',
                    }
                    for t in converted_ticket_files if t.file
                ]
            
            elif converted.ticket_type in ["pdf_template", "pdf_template_random"]:
                converted_pdf_tickets = [
                    {
                        'file_url': f"{domain}{t.file.url}" if t.file and t.file.url.startswith('/') else (t.file.url if t.file else ''),
                        'ticket_code': t.ticket_code,
                        'filename': t.file.name.split('/')[-1] if t.file else '',
                    }
                    for t in converted_ticket_files if t.file
                ]
            
            elif converted.ticket_type == "booking_code" and converted.generic_booking_code:
                converted_booking_code = converted.generic_booking_code
            
            elif converted.ticket_type == "instructions" and converted.ticket_instructions:
                converted_entry_instructions = converted.ticket_instructions
            
            elif converted.ticket_type == "box_office":
                converted_box_office_notes = converted.box_office_notes or ""

            # Calculate total tickets count for converted booking
            converted_total_tickets_count = (
                len(converted_ticket_codes_list) + 
                len(converted_pdf_tickets) + 
                len(converted_qr_tickets) + 
                (1 if converted_booking_code else 0) + 
                (1 if converted_entry_instructions else 0) + 
                (1 if converted.ticket_type == "box_office" else 0)
            )

            # Determine which ticket type is active and set all_tickets to that
            converted_all_tickets = []
            if converted_ticket_codes_list:
                converted_all_tickets = [{"type": "code", "value": code} for code in converted_ticket_codes_list]
            elif converted_pdf_tickets:
                converted_all_tickets = converted_pdf_tickets
            elif converted_qr_tickets:
                converted_all_tickets = converted_qr_tickets
            elif converted_booking_code:
                converted_all_tickets = [{"type": "booking_code", "code": converted_booking_code}]
            elif converted_entry_instructions:
                converted_all_tickets = [{"type": "instructions", "text": converted_entry_instructions}]
            elif converted.ticket_type == "box_office":
                converted_all_tickets = [{"type": "box_office", "notes": converted_box_office_notes}]

            # Format PDF tickets as HTML string for converted booking
            converted_pdf_tickets_html = ""
            if converted_pdf_tickets:
                converted_pdf_tickets_html = "<div class='pdf-tickets'>"
                for i, pdf in enumerate(converted_pdf_tickets, 1):
                    converted_pdf_tickets_html += f"""
                    <div class='ticket' style='margin-bottom: 15px; padding: 10px; border: 1px solid #ddd; border-radius: 4px;'>
                        <h4 style='margin-top: 0; color: #002663;'>PDF Ticket {i}</h4>
                        <p><a href='{pdf['file_url']}' style='background-color: #002663; color: white; padding: 8px 15px; text-decoration: none; border-radius: 4px; display: inline-block;'>Download {pdf['filename']}</a></p>
                        {f"<p><strong>Code:</strong> {pdf['ticket_code']}</p>" if pdf['ticket_code'] else ""}
                    </div>
                    """
                converted_pdf_tickets_html += "</div>"
            
            # Format ticket codes as HTML string for converted booking
            converted_ticket_codes_html = ""
            if converted_ticket_codes_list:
                converted_ticket_codes_html = "<div class='code-tickets'><ul style='list-style-type: none; padding: 0;'>"
                for code in converted_ticket_codes_list:
                    converted_ticket_codes_html += f"<li style='margin-bottom: 5px; padding: 8px; background-color: #f5f5f5; border-radius: 4px;'><strong>Code:</strong> {code}</li>"
                converted_ticket_codes_html += "</ul></div>"
            
            # Format QR tickets as HTML string for converted booking
            converted_qr_tickets_html = ""
            if converted_qr_tickets:
                converted_qr_tickets_html = "<div class='qr-tickets'>"
                for i, qr in enumerate(converted_qr_tickets, 1):
                    converted_qr_tickets_html += f"""
                    <div class='ticket' style='margin-bottom: 15px; padding: 10px; border: 1px solid #ddd; border-radius: 4px;'>
                        <h4 style='margin-top: 0; color: #002663;'>QR Ticket {i}</h4>
                        {f"<p><strong>QR Value:</strong> {qr['qr_value']}</p>" if qr['qr_value'] else ""}
                        <p><img src='{qr['file_url']}' alt='QR Code' style='max-width: 200px; border: 1px solid #ddd; padding: 5px;'></p>
                        <p><small>Filename: {qr['filename']}</small></p>
                    </div>
                    """
                converted_qr_tickets_html += "</div>"
            
            # Format all_tickets as HTML string for converted booking
            converted_all_tickets_html = ""
            if converted_all_tickets:
                converted_all_tickets_html = "<div class='all-tickets'>"
                for i, ticket in enumerate(converted_all_tickets, 1):
                    # Check if it's a code ticket
                    if ticket.get('type') == 'code' or (isinstance(ticket, dict) and 'value' in ticket and not ticket.get('file_url')):
                        value = ticket.get('value', '')
                        converted_all_tickets_html += f"""
                        <div class='ticket' style='margin-bottom: 15px; padding: 10px; border: 1px solid #ddd; border-radius: 4px;'>
                            <h4 style='margin-top: 0; color: #002663;'>Code Ticket {i}</h4>
                            <p><strong>Code:</strong> {value}</p>
                        </div>
                        """
                    
                    # Check if it's a QR ticket
                    elif ticket.get('type') == 'qr' or ticket.get('qr_value') is not None:
                        qr_value = ticket.get('qr_value', '')
                        file_url = ticket.get('file_url', '')
                        filename = ticket.get('filename', f'qr-code-{i}.png')
                        
                        converted_all_tickets_html += f"""
                        <div class='ticket' style='margin-bottom: 15px; padding: 10px; border: 1px solid #ddd; border-radius: 4px;'>
                            <h4 style='margin-top: 0; color: #002663;'>QR Ticket {i}</h4>
                            {f"<p><strong>QR Value:</strong> {qr_value}</p>" if qr_value else ""}
                            {f"<p><img src='{file_url}' alt='QR Code' style='max-width: 200px; border: 1px solid #ddd; padding: 5px;'></p>" if file_url else ""}
                            <p><small>Filename: {filename}</small></p>
                        </div>
                        """
                    
                    # Check if it's a PDF ticket
                    elif ticket.get('file_url') and (ticket.get('filename', '').lower().endswith('.pdf') or ticket.get('type') == 'pdf'):
                        file_url = ticket.get('file_url', '')
                        filename = ticket.get('filename', f'pdf-ticket-{i}.pdf')
                        ticket_code = ticket.get('ticket_code', '')
                        
                        converted_all_tickets_html += f"""
                        <div class='ticket' style='margin-bottom: 15px; padding: 10px; border: 1px solid #ddd; border-radius: 4px;'>
                            <h4 style='margin-top: 0; color: #002663;'>PDF Ticket {i}</h4>
                            <p><a href='{file_url}' style='background-color: #002663; color: white; padding: 8px 15px; text-decoration: none; border-radius: 4px; display: inline-block;'>Download {filename}</a></p>
                            {f"<p><strong>Code:</strong> {ticket_code}</p>" if ticket_code else ""}
                        </div>
                        """
                    
                    # Check if it's a file ticket
                    elif ticket.get('file_url'):
                        file_url = ticket.get('file_url', '')
                        filename = ticket.get('filename', f'file-{i}')
                        ticket_code = ticket.get('ticket_code', '')
                        
                        converted_all_tickets_html += f"""
                        <div class='ticket' style='margin-bottom: 15px; padding: 10px; border: 1px solid #ddd; border-radius: 4px;'>
                            <h4 style='margin-top: 0; color: #002663;'>Ticket {i}</h4>
                            <p><a href='{file_url}' style='background-color: #002663; color: white; padding: 8px 15px; text-decoration: none; border-radius: 4px; display: inline-block;'>Download {filename}</a></p>
                            {f"<p><strong>Code:</strong> {ticket_code}</p>" if ticket_code else ""}
                        </div>
                        """
                    
                    # Check if it's a booking code
                    elif ticket.get('type') == 'booking_code' or ticket.get('code'):
                        code = ticket.get('code', ticket.get('value', ''))
                        converted_all_tickets_html += f"""
                        <div class='ticket' style='margin-bottom: 15px; padding: 15px; background-color: #f5f5f5; border-radius: 4px; text-align: center;'>
                            <h4 style='margin-top: 0; color: #002663;'>Booking Code</h4>
                            <p style='font-size: 18px; font-weight: bold; color: #002663;'>{code}</p>
                        </div>
                        """
                    
                    # Check if it's instructions
                    elif ticket.get('type') == 'instructions' or ticket.get('text'):
                        text = ticket.get('text', ticket.get('value', ''))
                        converted_all_tickets_html += f"""
                        <div class='ticket' style='margin-bottom: 15px; padding: 15px; background-color: #f5f5f5; border-radius: 4px;'>
                            <h4 style='margin-top: 0; color: #002663;'>Entry Instructions</h4>
                            <p>{text}</p>
                        </div>
                        """
                    
                    # Check if it's box office
                    elif ticket.get('type') == 'box_office' or ticket.get('notes'):
                        notes = ticket.get('notes', ticket.get('value', ''))
                        converted_all_tickets_html += f"""
                        <div class='ticket' style='margin-bottom: 15px; padding: 15px; background-color: #f5f5f5; border-radius: 4px;'>
                            <h4 style='margin-top: 0; color: #002663;'>Box Office Collection</h4>
                            <p>{notes if notes else 'Please collect your tickets at the box office.'}</p>
                        </div>
                        """
                    
                    else:
                        converted_all_tickets_html += f"<div style='margin-bottom: 10px; padding: 8px; background-color: #f5f5f5; border-radius: 4px;'><pre>{ticket}</pre></div>"
                        
                converted_all_tickets_html += "</div>"
            
            # Format booking code as HTML for converted booking
            converted_booking_code_html = ""
            if converted_booking_code:
                converted_booking_code_html = f"<div style='margin-bottom: 15px; padding: 15px; background-color: #f5f5f5; border-radius: 4px; text-align: center;'><h4 style='margin-top: 0; color: #002663;'>Booking Code</h4><p style='font-size: 18px; font-weight: bold; color: #002663;'>{converted_booking_code}</p></div>"
            
            # Format entry instructions as HTML for converted booking
            converted_entry_instructions_html = ""
            if converted_entry_instructions:
                converted_entry_instructions_html = f"<div style='margin-bottom: 15px; padding: 15px; background-color: #f5f5f5; border-radius: 4px;'><h4 style='margin-top: 0; color: #002663;'>Entry Instructions</h4><p>{converted_entry_instructions}</p></div>"
            
            # Format box office notes as HTML for converted booking
            converted_box_office_html = ""
            if converted.ticket_type == "box_office":
                converted_box_office_html = f"<div style='margin-bottom: 15px; padding: 15px; background-color: #f5f5f5; border-radius: 4px;'><h4 style='margin-top: 0; color: #002663;'>Box Office Collection</h4><p>{converted_box_office_notes if converted_box_office_notes else 'Please collect your tickets at the box office.'}</p></div>"

            from django.utils.safestring import mark_safe
            context.update({
                
            #ticket (metadata)
            "ticket_type": converted.ticket_type,
            "ticket_sent_date": converted.ticket_sent_at.strftime("%d/%m/%Y %H:%M") if converted.ticket_sent_at else "",

            #all tickets
            "all_tickets": mark_safe(converted_all_tickets_html),
            "all_tickets_count": converted_total_tickets_count,

            #ticket codes
            "ticket_codes": mark_safe(converted_ticket_codes_html),
            "ticket_codes_count": len(converted_ticket_codes_list),

            #pdf tickets (file + code)
            "pdf_tickets": mark_safe(converted_pdf_tickets_html),
            "pdf_tickets_count": len(converted_pdf_tickets),

            #qr ticket (file + qr value)
            "qr_tickets": mark_safe(converted_qr_tickets_html),
            "qr_tickets_count": len(converted_qr_tickets),

            #booking code
            "booking_code": mark_safe(converted_booking_code_html),
        
            #entry instruction
            "entry_instructions": mark_safe(converted_entry_instructions_html),

            #box office collection
            "box_office_instructions": mark_safe(converted_box_office_html),
            })

    context.update(kwargs)

    if context.get("winner_deadline") and isinstance(context["winner_deadline"], datetime.datetime):
        context["winner_deadline"] = context["winner_deadline"].strftime("%d/%m/%Y %H:%M")
        context["winner_deadline_days"] = 3

    return context

# Ticket management views and helpers used by staff to upload files, assign codes, preview ticket output, and delete ticket data
# Covers both normal bookings and accepted draw bookings across individual and bulk actions
# -----------------------------
# Ticket Upload Mechanism
# -----------------------------
# Builds the staff ticket upload page with filtering, grouping, counts, and pagination for bookings and draw entries
@staff_member_required
def ticket_upload(request):
    q = (request.GET.get("q") or "").strip()
    venue_id = (request.GET.get("venue") or "").strip()
    sort = (request.GET.get("sort") or "date_desc").strip()
    kind = (request.GET.get("kind") or "").strip()
    show_all = request.GET.get("show_all") == "1"
    today = timezone.localdate()

    converted_booking_ids = TicketDrawBooking.objects.filter(
        converted_booking__isnull=False
    ).values_list("converted_booking_id", flat=True)

    def get_booking_reference(booking):
        if not booking:
            return "—"
        return f"REF-{booking.id}"

    def is_ticketed(obj):
        if not obj:
            return False

        return bool(
            obj.ticket_sent
            or obj.ticket_code
            or getattr(obj, "ticket_qr_value", "")
            or getattr(obj, "generic_booking_code", "")
            or getattr(obj, "ticket_instructions", "")
            or obj.tickets.exists()
            or obj.ticket_type in {
                "box_office",
                "codes",
                "pdf_template",
                "pdf_template_random",
                "qr_individual",
                "instructions",
                "booking_code",
            }
        )

    venues = []

    for a in Attraction.objects.order_by("name"):
        qs = Booking.objects.filter(
            attraction=a,
            cancelled=False,
            slot__date__gte=today,
        ).exclude(id__in=converted_booking_ids)

        tu_total = 0
        tu_ticketed = 0

        for b in qs:
            n = getattr(b, "num_tickets", 1) or 1
            tu_total += n
            if is_ticketed(b):
                tu_ticketed += n

        venues.append({
            "id": f"a-{a.id}",
            "name": a.name,
            "kind": "attraction",
            "tu_ticketed": tu_ticketed,
            "tu_total": tu_total,
        })

    for d in TicketDraw.objects.order_by("name"):
        qs = TicketDrawBooking.objects.filter(
            ticket_draw=d,
            cancelled=False,
            is_accepted=True,
            converted_booking__isnull=False,
            slot__date__gte=today,
        ).select_related("converted_booking")

        tu_total = 0
        tu_ticketed = 0

        for entry in qs:
            n = getattr(entry, "num_tickets", 1) or 1
            tu_total += n
            if is_ticketed(entry.converted_booking):
                tu_ticketed += n

        venues.append({
            "id": f"d-{d.id}",
            "name": f"{d.name} (Draw)",
            "kind": "draw",
            "tu_ticketed": tu_ticketed,
            "tu_total": tu_total,
        })

    attraction_qs = (
        Booking.objects
        .select_related("user", "slot", "attraction")
        .filter(cancelled=False)
        .exclude(id__in=converted_booking_ids)
    )

    draw_qs = (
        TicketDrawBooking.objects
        .select_related(
            "user",
            "ticket_draw",
            "slot",
            "converted_booking",
            "converted_booking__user",
            "converted_booking__slot",
            "converted_booking__attraction",
        )
        .filter(
            cancelled=False,
            is_accepted=True,
            converted_booking__isnull=False,
        )
    )

    if q:
        attraction_qs = attraction_qs.filter(
            Q(attraction__name__icontains=q)
            | Q(user__first_name__icontains=q)
            | Q(user__last_name__icontains=q)
            | Q(full_name__icontains=q)
            | Q(email__icontains=q)
            | Q(id__icontains=q)
            | Q(ticket_code__icontains=q)
            | Q(generic_booking_code__icontains=q)
        )

        draw_qs = draw_qs.filter(
            Q(ticket_draw__name__icontains=q)
            | Q(user__first_name__icontains=q)
            | Q(user__last_name__icontains=q)
            | Q(full_name__icontains=q)
            | Q(email__icontains=q)
            | Q(id__icontains=q)
            | Q(converted_booking__ticket_code__icontains=q)
            | Q(converted_booking__generic_booking_code__icontains=q)
        )

    if venue_id:
        if venue_id.startswith("a-"):
            attraction_id = venue_id.split("-", 1)[1]
            attraction_qs = attraction_qs.filter(attraction_id=attraction_id)
            draw_qs = draw_qs.none()
        elif venue_id.startswith("d-"):
            draw_id = venue_id.split("-", 1)[1]
            draw_qs = draw_qs.filter(ticket_draw_id=draw_id)
            attraction_qs = attraction_qs.none()

    if kind == "draw":
        attraction_qs = attraction_qs.none()
    elif kind == "attraction":
        draw_qs = draw_qs.none()

    booking_rows = []
    for b in attraction_qs:
        ticketed = is_ticketed(b)

        booking_rows.append({
            "group_key": f"a-{b.attraction_id}",
            "row_id": b.id,
            "id": b.id,
            "row_kind": "b",
            "venue_name": b.attraction.name if b.attraction else "",
            "first_name": b.user.first_name if b.user else (b.full_name.split(" ")[0] if b.full_name else ""),
            "last_name": b.user.last_name if b.user else (" ".join(b.full_name.split(" ")[1:]) if b.full_name and " " in b.full_name else ""),
            "guid": b.user.profile.staff_guid if b.user and hasattr(b.user, "profile") else "",
            "booking_date": b.slot.date.strftime("%d/%m/%Y") if b.slot and b.slot.date else "",
            "booking_reference": get_booking_reference(b),
            "sort_date": b.slot.date if b.slot else None,
            "sort_time": b.slot.time if b.slot else None,
            "ticket_sent": ticketed,
            "ticket_type": b.ticket_type if b.ticket_type else "—",
            "ticket_code": b.ticket_code,
            "ticket_file_url": b.ticket_file.url if getattr(b, "ticket_file", None) else "",
            "num_tickets": b.num_tickets if hasattr(b, "num_tickets") else 1,
            "is_draw": False,
            "is_past": bool(b.slot and b.slot.date < today),
        })

    draw_rows = []
    for d in draw_qs:
        b = d.converted_booking
        ticketed = is_ticketed(b)

        draw_rows.append({
            "group_key": f"d-{d.ticket_draw_id}",
            "row_id": d.id,
            "id": d.id,
            "row_kind": "d",
            "venue_name": f"{d.ticket_draw.name} (Draw)" if d.ticket_draw else "",
            "first_name": d.user.first_name if d.user else "",
            "last_name": d.user.last_name if d.user else "",
            "guid": d.user.profile.staff_guid if d.user and hasattr(d.user, "profile") else "",
            "booking_date": d.slot.date.strftime("%d/%m/%Y") if d.slot and d.slot.date else "",
            "booking_reference": get_booking_reference(b),
            "sort_date": d.slot.date if d.slot else None,
            "sort_time": d.slot.time if d.slot else None,
            "ticket_sent": ticketed,
            "ticket_type": b.ticket_type if b and b.ticket_type else "—",
            "ticket_code": b.ticket_code if b else "",
            "ticket_file_url": b.ticket_file.url if b and getattr(b, "ticket_file", None) else "",
            "num_tickets": d.num_tickets if hasattr(d, "num_tickets") else (
                b.num_tickets if b and hasattr(b, "num_tickets") else 1
            ),
            "is_draw": True,
            "is_past": bool(d.slot and d.slot.date < today),
        })

    raw_rows = booking_rows + draw_rows

    grouped_counts = {}
    for row in raw_rows:
        if row.get("is_past"):
            continue

        key = row["group_key"]
        grouped_counts.setdefault(key, {"ticketed": 0, "total": 0})

        tickets_for_row = row.get("num_tickets", 1) or 1
        grouped_counts[key]["total"] += tickets_for_row

        if row["ticket_sent"]:
            grouped_counts[key]["ticketed"] += tickets_for_row

    all_rows = []
    for row in raw_rows:
        if kind == "ticketed" and not row["ticket_sent"]:
            continue
        if kind == "unticketed" and row["ticket_sent"]:
            continue

        counts = grouped_counts.get(row["group_key"], {"ticketed": 0, "total": 0})
        row["ticketed_count"] = counts["ticketed"]
        row["total_count"] = counts["total"]
        all_rows.append(row)

    if sort == "date_asc":
        all_rows.sort(key=lambda r: (r["sort_date"] or "", r["sort_time"] or ""))
    elif sort == "surname":
        all_rows.sort(key=lambda r: (r["last_name"] or "", r["first_name"] or ""))
    elif sort == "venue":
        all_rows.sort(key=lambda r: (r["venue_name"] or "", r["sort_date"] or ""))
    else:
        all_rows.sort(key=lambda r: (r["sort_date"] or "", r["sort_time"] or ""), reverse=True)

    paginator = Paginator(all_rows, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    paged_rows = list(page_obj.object_list)
    rows = [r for r in paged_rows if not r.get("is_draw")]
    draw_rows = [r for r in paged_rows if r.get("is_draw")]
    empty_rows = [] if show_all else [None] * max(0, 10 - len(paged_rows))

    venue_ticketed = 0
    venue_total = 0

    if venue_id.startswith("a-"):
        attraction_id = venue_id.split("-", 1)[1]

        venue_bookings = Booking.objects.filter(
            attraction_id=attraction_id,
            cancelled=False,
            slot__date__gte=today,
        ).exclude(id__in=converted_booking_ids)

        for b in venue_bookings:
            tickets_for_booking = getattr(b, "num_tickets", 1) or 1
            venue_total += tickets_for_booking
            if is_ticketed(b):
                venue_ticketed += tickets_for_booking

    elif venue_id.startswith("d-"):
        draw_id = venue_id.split("-", 1)[1]

        venue_draws = TicketDrawBooking.objects.filter(
            ticket_draw_id=draw_id,
            cancelled=False,
            is_accepted=True,
            converted_booking__isnull=False,
            slot__date__gte=today,
        ).select_related("converted_booking")

        for d in venue_draws:
            tickets_for_booking = getattr(d, "num_tickets", 1) or 1
            venue_total += tickets_for_booking
            if is_ticketed(d.converted_booking):
                venue_ticketed += tickets_for_booking

    return render(
        request,
        "fergusonbequest/ticket_upload.html",
        {
            "venues": venues,
            "rows": rows,
            "draw_rows": draw_rows,
            "empty_rows": empty_rows,
            "page_obj": page_obj,
            "tickets": page_obj.object_list,
            "q": q,
            "venue_id": venue_id,
            "sort": sort,
            "kind": kind,
            "show_all": show_all,
            "ticket_view_url_template": reverse("ticket_view", args=["999999"]),
            "venue_ticketed": venue_ticketed,
            "venue_total": venue_total,
        },
    )
# Returns all uploaded ticket file URLs for a booking or accepted draw booking as JSON
@login_required
def ticket_list(request, booking_id):
    """
    Returns a list of ticket URLs for a specific booking.
    Supports both standard Bookings and TicketDrawBookings.
    """
    raw_id = str(booking_id)

    if ":" in raw_id:
        row_kind, raw_pk = raw_id.split(":", 1)
    else:
        row_kind, raw_pk = "b", raw_id

    booking = None

    if row_kind == "d":
        draw_entry = (
            TicketDrawBooking.objects
            .filter(id=raw_pk)
            .select_related("converted_booking")
            .first()
        )
        if draw_entry:
            booking = draw_entry.converted_booking
    else:
        booking = Booking.objects.filter(id=raw_pk).first()

    if not booking or (hasattr(booking, "cancelled") and booking.cancelled):
        return JsonResponse({"tickets": []})

    ref = f"REF-{booking.id}"

    tickets = [
        {
            "url": request.build_absolute_uri(t.file.url),
            "qr_value": t.qr_value or "",
            "ticket_code": t.ticket_code or "",
        }
        for t in booking.tickets.all()
        if t.file
    ]

    return JsonResponse({
        "tickets": tickets,
        "booking_reference": ref,
        "ticket_type": booking.ticket_type,
    })

# Marks a booking as ticket sent from the staff upload screen
@staff_member_required
@require_http_methods(["POST"])
def ticket_upload_send(request):
    booking_id = request.POST.get("row_id")
    booking = get_object_or_404(Booking, id=booking_id)

    if not booking.ticket_type:
        messages.error(request, "No ticket configured for this booking.")
        return redirect("ticket_upload")

    booking.ticket_sent = True
    booking.ticket_sent_at = timezone.now()
    booking.save(update_fields=["ticket_sent", "ticket_sent_at"])

    messages.success(request, f"Ticket marked as sent for booking #{booking.id}.")
    return redirect("ticket_upload")

# Redirects back to the main ticket upload page when viewing all rows
@staff_member_required
@require_http_methods(["POST"])
def ticket_upload_view_all(request):
    return redirect("ticket_upload")

# Builds the best available ticket response for a booking
def _ticket_response_for_booking(booking):
    ref = booking.generic_booking_code or f"REF-{booking.id}"
    first_ticket = booking.tickets.first()

    if first_ticket and first_ticket.file:
        try:
            return FileResponse(
                first_ticket.file.open("rb"),
                as_attachment=True,
                filename=first_ticket.file.name.split("/")[-1],
            )
        except Exception:
            pass

    if booking.ticket_type == "qr_individual":
        text = "QR ticket"
        if booking.ticket_qr_value:
            text += f"\n\nQR Value: {booking.ticket_qr_value}"

        return HttpResponse(
            text,
            content_type="text/plain",
        )

    if booking.ticket_type == "instructions" and booking.ticket_instructions:
        return HttpResponse(
            f"Instructions:\n{booking.ticket_instructions}",
            content_type="text/plain",
        )

    if booking.ticket_type == "box_office":
        text = "Collect tickets at the box office (no digital ticket)."
        if booking.box_office_notes:
            text += f"\n\n{booking.box_office_notes}"

        return HttpResponse(
            text,
            content_type="text/plain",
        )

    if booking.ticket_type == "codes":
        return HttpResponse(
            f"Ticket Code: {booking.ticket_code}",
            content_type="text/plain",
        )

    if booking.ticket_type == "booking_code" and booking.generic_booking_code:
        return HttpResponse(
            f"Booking Code: {booking.generic_booking_code}",
            content_type="text/plain",
        )

    return HttpResponse(
        "No ticket info available for this booking.",
        status=404,
    )
# Lets staff preview ticket output for either a normal booking or a draw booking
@staff_member_required
def ticket_view(request, booking_id):
    raw_id = str(booking_id)

    if ":" in raw_id:
        row_kind, raw_pk = raw_id.split(":", 1)
    else:
        row_kind, raw_pk = "b", raw_id

    if row_kind == "d":
        draw_booking = (
            TicketDrawBooking.objects
            .select_related("converted_booking")
            .filter(id=raw_pk)
            .first()
        )

        if not draw_booking:
            return HttpResponse("Ticket not found.", status=404)

        if draw_booking.converted_booking:
            return _ticket_response_for_booking(draw_booking.converted_booking)

        return HttpResponse("No ticket info available for this booking.", status=404)

    booking = Booking.objects.filter(id=raw_pk).first()
    if booking:
        return _ticket_response_for_booking(booking)

    return HttpResponse("Ticket not found.", status=404)

# Lets a logged-in user view their own ticket, while staff can view any ticket
@login_required
def user_ticket_view(request, booking_id):

    booking = Booking.objects.filter(id=booking_id, cancelled=False).first()

    if booking:
        if booking.user_id != request.user.id and not request.user.is_staff:
            return HttpResponseForbidden("Not allowed.")
        return _ticket_response_for_booking(booking)

    # If not a normal booking try a draw booking
    draw_booking = TicketDrawBooking.objects.select_related("converted_booking").filter(
        id=booking_id,
        cancelled=False,
    ).first()

    if draw_booking:
        if draw_booking.user_id != request.user.id and not request.user.is_staff:
            return HttpResponseForbidden("Not allowed.")

        if draw_booking.converted_booking:
            return _ticket_response_for_booking(draw_booking.converted_booking)

        return HttpResponse("No ticket info available for this booking.", status=404)

    return HttpResponse("Ticket not found.", status=404)


# Applies a shared ticket assignment action to multiple bookings and marks them as sent
def _bulk_assign(request, bookings, *, apply_fn, label):
    """
    apply_fn(booking, idx) should set fields on booking and return update_fields list.
    """
    total = len(bookings)
    assigned = 0
    now = timezone.now()

    for idx, b in enumerate(bookings):
        update_fields = apply_fn(b, idx)

        b.ticket_sent = True
        b.ticket_sent_at = now

        if "ticket_sent" not in update_fields:
            update_fields.append("ticket_sent")
        if "ticket_sent_at" not in update_fields:
            update_fields.append("ticket_sent_at")

        b.save(update_fields=update_fields)
        assigned += 1

    messages.success(request, f"{label}: assigned {assigned} of {total}.")
    return assigned


# Distributes tickets in bulk to all unsent bookings for a selected venue using the chosen ticket method
@staff_member_required
@require_POST
def venue_distribute_tickets(request):
    venue_id = (request.POST.get("venue_id") or "").strip()
    ticket_type = (request.POST.get("ticket_type") or "").strip()

    is_draw_venue = venue_id.startswith("d-")

    if venue_id.startswith("a-"):
        venue_id = venue_id.split("-", 1)[1]
    elif venue_id.startswith("d-"):
        venue_id = venue_id.split("-", 1)[1]

    try:
        venue_id = int(venue_id)
    except ValueError:
        messages.error(request, "Invalid venue selected.")
        return redirect("ticket_upload")

    # Inputs
    codes_file = request.FILES.get("codes_file")
    codes_text = request.POST.get("codes_text")
    trim_spaces = bool(request.POST.get("trim_spaces"))
    dedupe_codes = bool(request.POST.get("dedupe_codes"))

    if not venue_id:
        messages.error(request, "No venue selected.")
        return redirect("/ticket-upload/?open_upload=1")

    if not ticket_type:
        messages.error(request, "No ticket type selected.")
        return redirect("ticket_upload")

    def parse_codes():
        codes = []

        if codes_file:
            wrapper = TextIOWrapper(codes_file, encoding="utf-8")
            for line in wrapper:
                codes.append(line.rstrip("\n"))
        elif codes_text:
            codes = codes_text.splitlines()

        if trim_spaces:
            codes = [c.strip() for c in codes]

        codes = [c for c in codes if c]

        if codes and codes[0].strip().lower() in {"ticket_code", "code", "codes"}:
            codes = codes[1:]

        if dedupe_codes:
            seen = set()
            unique = []
            for c in codes:
                if c not in seen:
                    unique.append(c)
                    seen.add(c)
            codes = unique

        return codes

    with transaction.atomic():
        if is_draw_venue:
            # For draw venues, get the draw bookings
            draw_bookings = list(
                TicketDrawBooking.objects.filter(
                    ticket_draw_id=venue_id,
                    cancelled=False,
                    is_accepted=True,
                    converted_booking__isnull=False,
                    converted_booking__ticket_sent=False,
                    slot__date__gte=timezone.now().date(),
                )
                .select_for_update()
                .select_related('converted_booking')
                .order_by("created_at")
            )
            
            bookings = [db.converted_booking for db in draw_bookings]
            
            if not bookings:
                messages.info(request, "No unsent draw bookings found for this venue.")
                return redirect("ticket_upload")
        else:
            # lock only unsent bookings for this venue
            bookings = list(
                Booking.objects.filter(
                    attraction_id=venue_id,
                    cancelled=False,
                    ticket_sent=False,
                    slot__date__gte=timezone.now().date(),
                )
                .exclude(ticket_code__isnull=False).exclude(ticket_code="")
                .select_for_update()
                .order_by("created_at")
            )

        if not bookings:
            messages.info(request, "No unsent bookings found for this venue.")
            return redirect("ticket_upload")

        total_needed = sum(max(1, b.num_tickets or 1) for b in bookings)

        # E-ticket codes  (applies to ALL unsent)
        if ticket_type == "codes":
            codes = parse_codes()

            if not codes:
                messages.error(request, "No codes provided.")
                return redirect("ticket_upload")

            to_assign = min(len(codes), total_needed)
            extra = max(0, len(codes) - to_assign)

            def apply_fn(b, idx):
                b.ticket_type = "codes"
                b.ticket_code = codes[idx]
                return ["ticket_type", "ticket_code"]

            _bulk_assign(request, bookings[:to_assign], apply_fn=apply_fn, label="E-ticket codes")

            if is_draw_venue and draw_bookings:
                for db in draw_bookings[:to_assign]:
                    send_draw_booking_email_ticket_distribution(db)

            remaining = total_needed - to_assign
            if remaining:
                messages.info(request, f"{remaining} booking(s) still need tickets.")
            if extra:
                messages.info(request, f"{extra} extra code(s) ignored.")

            return redirect("ticket_upload")

        # PDF upload (applies to ALL unsent)
        if ticket_type in {"pdf_template", "pdf_template_random"}:
            files = request.FILES.getlist("ticket_files")

            if not files:
                pdf = request.FILES.get("ticket_file")
                files = [pdf] if pdf else []

            if not files:
                messages.error(request, "Please upload at least one PDF.")
                return redirect("ticket_upload")

            generate_codes = bool(request.POST.get("pdf_generate_codes"))
            raw_codes = (request.POST.get("pdf_codes_bulk") or "").strip()
            manual_codes = [line.strip() for line in raw_codes.splitlines() if line.strip()]

            if not generate_codes and manual_codes and len(manual_codes) != len(files):
                messages.error(
                    request,
                    f"You uploaded {len(files)} PDF file(s) but entered {len(manual_codes)} code(s). "
                    "Please provide one code per PDF, in the same order."
                )
                return redirect("ticket_upload")

            used = set(
                BookingTicket.objects.exclude(ticket_code__isnull=True)
                .exclude(ticket_code="")
                .values_list("ticket_code", flat=True)
            )

            def generate_random_code():
                while True:
                    c = "FB-" + "".join(random.choices(string.ascii_uppercase + string.digits, k=8))
                    if c not in used:
                        used.add(c)
                        return c

            to_assign = min(len(files), total_needed)
            extra = max(0, len(files) - to_assign)

            def apply_fn(b, idx):
                b.ticket_type = "pdf_template"
                b.ticket_code = ""

                per_file_code = generate_random_code() if generate_codes else (
                    manual_codes[idx] if idx < len(manual_codes) else "")

                BookingTicket.objects.create(
                    booking=b,
                    file=files[idx],
                    ticket_code=per_file_code,
                    sort_order=idx,
                )
                return ["ticket_type", "ticket_code"]

            _bulk_assign(request, bookings[:to_assign], apply_fn=apply_fn, label="PDF tickets")

            if is_draw_venue and draw_bookings:
                for db in draw_bookings[:to_assign]:
                    send_draw_booking_email_ticket_distribution(db)

            remaining = total_needed - to_assign
            if remaining:
                messages.info(request, f"{remaining} booking(s) still need tickets.")
            if extra:
                messages.info(request, f"{extra} extra file(s) ignored.")

            return redirect("ticket_upload")

        # QR bulk individual files (applies to ALL unsent)
        if ticket_type == "qr_individual":
            files = request.FILES.getlist("qr_files")
            if not files:
                messages.error(request, "Please upload at least one QR file.")
                return redirect("ticket_upload")

            raw_qr_values = (request.POST.get("ticket_qr_values_bulk") or "").strip()
            qr_values = [line.strip() for line in raw_qr_values.splitlines() if line.strip()]

            if qr_values and len(qr_values) != len(files):
                messages.error(
                    request,
                    f"You uploaded {len(files)} QR file(s) but entered {len(qr_values)} QR value(s). "
                    "Please provide one QR value per file, in the same order."
                )
                return redirect("ticket_upload")

            to_assign = min(len(files), total_needed)
            extra = max(0, len(files) - to_assign)

            def apply_fn(b, idx):
                b.ticket_type = "qr_individual"
                b.ticket_code = ""
                BookingTicket.objects.create(
                    booking=b,
                    file=files[idx],
                    qr_value=qr_values[idx] if idx < len(qr_values) else "",
                    sort_order=idx,
                )
                return ["ticket_type", "ticket_code"]

            _bulk_assign(request, bookings[:to_assign], apply_fn=apply_fn, label="QR tickets")
            if is_draw_venue and draw_bookings:
                for db in draw_bookings[:to_assign]:
                    send_draw_booking_email_ticket_distribution(db)

            remaining = total_needed - to_assign
            if remaining:
                messages.info(request, f"{remaining} booking(s) still need tickets.")
            if extra:
                messages.info(request, f"{extra} extra file(s) ignored.")
            return redirect("ticket_upload")

        if ticket_type == "booking_code":
            code = (request.POST.get("booking_code_bulk") or "").strip()
            if not code:
                messages.error(request, "Please enter a booking code.")
                return redirect("ticket_upload")

            def apply_fn(b, idx):
                b.ticket_type = "booking_code"
                b.generic_booking_code = code
                b.ticket_code = ""
                return ["ticket_type", "generic_booking_code", "ticket_code"]

            _bulk_assign(request, bookings, apply_fn=apply_fn, label="Booking code")

            if is_draw_venue and draw_bookings:
                for db in draw_bookings:
                    send_draw_booking_email_ticket_distribution(db)

            return redirect("ticket_upload")

        # instructions (applies to ALL unsent)
        if ticket_type == "instructions":
            instructions = (request.POST.get("instructions") or "").strip()
            if not instructions:
                messages.error(request, "Please enter instructions.")
                return redirect("ticket_upload")

            def apply_fn(b, idx):
                b.ticket_type = "instructions"
                b.ticket_code = instructions
                return ["ticket_type", "ticket_code"]

            _bulk_assign(request, bookings, apply_fn=apply_fn, label="Instructions")

            if is_draw_venue and draw_bookings:
                for db in draw_bookings:
                    send_draw_booking_email_ticket_distribution(db)

            return redirect("ticket_upload")

        # Box office (applies to ALL unsent)
        if ticket_type == "box_office":
            notes = (request.POST.get("box_office_notes") or "").strip()

            def apply_fn(b, idx):
                b.ticket_type = "box_office"
                b.box_office_notes = notes
                b.ticket_code = ""
                b.ticket_qr_value = ""
                b.ticket_instructions = ""
                b.generic_booking_code = ""
                b.tickets.all().delete()
                return [
                    "ticket_type",
                    "box_office_notes",
                    "ticket_code",
                    "ticket_qr_value",
                    "ticket_instructions",
                    "generic_booking_code",
                ]

            _bulk_assign(request, bookings, apply_fn=apply_fn, label="Box office collection")

            if is_draw_venue and draw_bookings:
                for db in draw_bookings:
                    send_draw_booking_email_ticket_distribution(db)

            return redirect("ticket_upload")

# Assigns or uploads ticket data for a single booking or accepted draw booking
@staff_member_required
@require_POST
def individual_booking(request):
    booking_id = (request.POST.get("booking_id") or "").strip()
    ticket_type = (request.POST.get("ticket_type") or "").strip()
    row_kind = request.POST.get("row_kind", "b")

    if not booking_id:
        messages.error(request, "Missing booking id.")
        return redirect("ticket_upload")

    if not ticket_type:
        messages.error(request, "No ticket type selected.")
        return redirect("ticket_upload")

    is_draw = False
    draw_booking = None

    if row_kind == "d":
        draw_booking = get_object_or_404(TicketDrawBooking, id=booking_id, cancelled=False)
        is_draw = True

        if not draw_booking.converted_booking:
            messages.error(request, "This draw entry has not been accepted yet.")
            return redirect("ticket_upload")

        booking = draw_booking.converted_booking
    else:
        booking = get_object_or_404(Booking, id=booking_id, cancelled=False)


    template_pdf = request.FILES.get("ticket_file")
    codes_file = request.FILES.get("codes_file")
    codes_text = request.POST.get("codes_text")
    template_files = request.FILES.getlist("ticket_files")
    qr_files_individual = request.FILES.getlist("qr_files_individual")
    requested_num_tickets = int(request.POST.get("num_tickets") or getattr(booking, "num_tickets", 1) or 1)
    requested_num_tickets = max(1, requested_num_tickets)

    trim_spaces = bool(request.POST.get("trim_spaces"))
    dedupe_codes = bool(request.POST.get("dedupe_codes"))

    now = timezone.now()
    visit_datetime = datetime.combine(booking.slot.date, datetime.min.time())
    if timezone.is_aware(timezone.now()):
        visit_datetime = timezone.make_aware(visit_datetime)

    ticket_visible_at = visit_datetime - timedelta(days=3)

    # Booking code (generic)
    if ticket_type == "booking_code":
        code = (request.POST.get("booking_code") or "").strip()
        if not code:
            messages.error(request, "Please enter a booking code.")
            return redirect("ticket_upload")

        booking.ticket_type = "booking_code"
        booking.generic_booking_code = code
        booking.ticket_visible_at = ticket_visible_at
        booking.ticket_code = ""
        booking.ticket_sent = True
        booking.ticket_sent_at = now
        booking.save(update_fields=[
            "ticket_type", "generic_booking_code", "ticket_visible_at",
            "ticket_code", "ticket_sent", "ticket_sent_at"
        ])

        if is_draw and draw_booking:
            send_draw_booking_email_ticket_distribution(draw_booking)

        messages.success(request, f"Booking code set for booking #{booking.id}.")
        return redirect("ticket_upload")

    # Staff-card instructions
    if ticket_type == "instructions":
        text = (request.POST.get("instructions") or "").strip()
        if not text:
            messages.error(request, "Please enter instructions.")
            return redirect("ticket_upload")

        booking.ticket_type = "instructions"
        booking.ticket_instructions = text
        booking.ticket_code = ""
        booking.ticket_sent = True
        booking.ticket_sent_at = now
        booking.save(update_fields=[
            "ticket_type", "ticket_instructions", "ticket_code",
            "ticket_sent", "ticket_sent_at"
        ])

        if is_draw and draw_booking:
            send_draw_booking_email_ticket_distribution(draw_booking)

        messages.success(request, f"Instructions set for booking #{booking.id}.")
        return redirect("ticket_upload")
    # qr individually sent
    if ticket_type == "qr_individual":
        files = qr_files_individual or ([] if not request.FILES.get("qr_file") else [request.FILES.get("qr_file")])

        if not files:
            messages.error(request, "Please upload at least one QR ticket file.")
            return redirect("ticket_upload")

        if len(files) != requested_num_tickets:
            messages.error(
                request,
                f"This booking requires {requested_num_tickets} ticket(s). "
                f"You uploaded {len(files)} file(s). Please upload exactly {requested_num_tickets}."
            )
            return redirect("ticket_upload")

        raw_qr_values = (request.POST.get("ticket_qr_values_individual") or "").strip()
        qr_values = [line.strip() for line in raw_qr_values.splitlines() if line.strip()]

        if qr_values and len(qr_values) != len(files):
            messages.error(
                request,
                f"You uploaded {len(files)} QR ticket file(s) but entered {len(qr_values)} QR value(s). "
                "Please enter one QR value per file."
            )
            return redirect("ticket_upload")

        booking.ticket_type = "qr_individual"
        booking.ticket_code = ""
        booking.ticket_visible_at = ticket_visible_at
        booking.ticket_sent = True
        booking.ticket_sent_at = now
        booking.save(update_fields=[
            "ticket_type", "ticket_code",
            "ticket_visible_at", "ticket_sent", "ticket_sent_at"
        ])

        booking.tickets.all().delete()
        for i, f in enumerate(files):
            BookingTicket.objects.create(
                booking=booking,
                file=f,
                qr_value=qr_values[i] if i < len(qr_values) else "",
                sort_order=i,
            )

        if is_draw and draw_booking:
            send_draw_booking_email_ticket_distribution(draw_booking)

        messages.success(request, f"{len(files)} QR ticket(s) uploaded for booking #{booking.id}.")
        return redirect("ticket_upload")

    # E-ticket codes (one per ticket)
    if ticket_type == "codes":
        codes = []

        # Individual modal textarea
        single_codes_text = (request.POST.get("single_codes_text") or "").strip()
        if single_codes_text:
            lines = single_codes_text.splitlines()
            codes = [l.strip() for l in lines if l.strip()]

        # Fallbacks for older inputs / bulk-style input
        elif codes_file:
            wrapper = TextIOWrapper(codes_file, encoding="utf-8")
            lines = [line.strip() for line in wrapper if line.strip()]
            codes = lines

        elif codes_text:
            lines = codes_text.splitlines()
            if trim_spaces:
                lines = [l.strip() for l in lines]
            lines = [l for l in lines if l]

            if dedupe_codes:
                seen = set()
                uniq = []
                for l in lines:
                    if l not in seen:
                        uniq.append(l)
                        seen.add(l)
                lines = uniq

            codes = lines

        if not codes:
            messages.error(request, "Please provide at least one code.")
            return redirect("ticket_upload")

        if len(codes) != requested_num_tickets:
            messages.error(
                request,
                f"This booking requires {requested_num_tickets} ticket(s). "
                f"You entered {len(codes)} code(s). Please enter exactly {requested_num_tickets}."
            )
            return redirect("ticket_upload")

        booking.ticket_type = "codes"
        booking.ticket_code = "\n".join(codes)
        booking.ticket_visible_at = ticket_visible_at
        booking.ticket_sent = True
        booking.ticket_sent_at = now

        # Clear fields from other ticket modes
        booking.ticket_instructions = ""
        booking.generic_booking_code = ""
        booking.ticket_qr_value = ""
        # Remove uploaded file tickets if switching from file-based mode
        booking.tickets.all().delete()
        booking.save(update_fields=[
            "ticket_type",
            "ticket_code",
            "ticket_visible_at",
            "ticket_sent",
            "ticket_sent_at",
            "ticket_instructions",
            "generic_booking_code",
            "ticket_qr_value",
        ])

        if is_draw and draw_booking:
            send_draw_booking_email_ticket_distribution(draw_booking)

        messages.success(request, f"{len(codes)} code(s) assigned for booking #{booking.id}.")
        return redirect("ticket_upload")

    # PDF template or template + random code
    if ticket_type in {"pdf_template", "pdf_template_random"}:
        files = template_files or ([] if not template_pdf else [template_pdf])

        if not files:
            messages.error(request, "Please upload at least one PDF/file.")
            return redirect("ticket_upload")

        if len(files) != requested_num_tickets:
            messages.error(
                request,
                f"This booking requires {requested_num_tickets} ticket(s). "
                f"You uploaded {len(files)} file(s). Please upload exactly {requested_num_tickets}."
            )
            return redirect("ticket_upload")

        generate_codes = bool(request.POST.get("pdf_generate_codes"))
        raw_codes = (request.POST.get("pdf_codes_individual") or "").strip()
        manual_codes = [line.strip() for line in raw_codes.splitlines() if line.strip()]

        if not generate_codes and manual_codes and len(manual_codes) != len(files):
            messages.error(
                request,
                f"You uploaded {len(files)} PDF file(s) but entered {len(manual_codes)} code(s). "
                "Please provide one code per PDF."
            )
            return redirect("ticket_upload")

        used = set(
            BookingTicket.objects.exclude(ticket_code__isnull=True)
            .exclude(ticket_code="")
            .values_list("ticket_code", flat=True)
        )

        def generate_random_code():
            while True:
                c = "FB-" + "".join(random.choices(string.ascii_uppercase + string.digits, k=8))
                if c not in used:
                    used.add(c)
                    return c

        booking.ticket_type = "pdf_template"
        booking.ticket_code = ""
        booking.ticket_visible_at = ticket_visible_at
        booking.ticket_sent = True
        booking.ticket_sent_at = now
        booking.save(update_fields=[
            "ticket_type", "ticket_code", "ticket_visible_at",
            "ticket_sent", "ticket_sent_at"
        ])

        booking.tickets.all().delete()

        for i, f in enumerate(files):
            per_file_code = generate_random_code() if generate_codes else (
                manual_codes[i] if i < len(manual_codes) else "")
            BookingTicket.objects.create(
                booking=booking,
                file=f,
                ticket_code=per_file_code,
                sort_order=i,
            )

        if is_draw and draw_booking:
            send_draw_booking_email_ticket_distribution(draw_booking)

        messages.success(request, f"{len(files)} PDF ticket(s) assigned for booking #{booking.id}.")
        return redirect("ticket_upload")
    
    if ticket_type == "box_office":
        notes = (request.POST.get("box_office_notes") or "").strip()

        booking.ticket_type = "box_office"
        booking.box_office_notes = notes
        booking.ticket_code = ""
        booking.ticket_qr_value = ""
        booking.ticket_instructions = ""
        booking.generic_booking_code = ""
        booking.ticket_visible_at = ticket_visible_at
        booking.ticket_sent = True
        booking.ticket_sent_at = now
        booking.tickets.all().delete()
        booking.save(update_fields=[
            "ticket_type",
            "box_office_notes",
            "ticket_code",
            "ticket_qr_value",
            "ticket_instructions",
            "generic_booking_code",
            "ticket_visible_at",
            "ticket_sent",
            "ticket_sent_at",
        ])

        if is_draw and draw_booking:
            send_draw_booking_email_ticket_distribution(draw_booking)

        messages.success(request, f"Box office collection set for booking #{booking.id}.")
        return redirect("ticket_upload")

# Clears irrelevant ticket fields so a booking stays consistent with its selected ticket type
def normalize_ticket_fields(booking):
    """
    Ensure only relevant ticket fields are populated for the chosen ticket_type.
    Call before booking.save() if you want to enforce consistency.
    """
    if booking.ticket_type == "box_office":
        booking.ticket_code = ""
        booking.ticket_qr_value = ""
        booking.ticket_instructions = ""

    elif booking.ticket_type == "codes":
        booking.ticket_qr_value = ""
        booking.ticket_instructions = ""

    elif booking.ticket_type in {"pdf_template", "pdf_template_random"}:
        if booking.ticket_type == "pdf_template":
            booking.ticket_code = ""
        booking.ticket_qr_value = ""
        booking.ticket_instructions = ""

    elif booking.ticket_type == "qr_individual":
        booking.ticket_code = ""
        booking.ticket_instructions = ""

    elif booking.ticket_type == "instructions":
        booking.ticket_code = ""
        booking.ticket_qr_value = ""

    elif booking.ticket_type == "booking_code":
        booking.ticket_code = ""
        booking.ticket_qr_value = ""
        booking.ticket_instructions = ""

# Removes ticket data and files from a single booking or draw-linked booking
@staff_member_required
@require_POST
def ticket_upload_delete(request):
    row_id = request.POST.get("row_id")
    row_kind = (request.POST.get("row_kind") or "b").strip().lower()

    if row_kind == "d":
        draw_booking = get_object_or_404(TicketDrawBooking, id=row_id)

        if not draw_booking.converted_booking:
            messages.error(request, "Draw booking has no ticket yet.")
            return redirect("ticket_upload")

        obj = draw_booking.converted_booking
    else:
        obj = get_object_or_404(Booking, id=row_id)

    obj.tickets.all().delete()
    obj.ticket_type = None
    obj.ticket_code = None
    obj.ticket_sent = False
    obj.ticket_sent_at = None
    obj.ticket_qr_value = ""
    obj.ticket_instructions = ""
    obj.generic_booking_code = ""

    obj.save()

    messages.success(request, f"Ticket removed for booking #{obj.id}.")
    return redirect("ticket_upload")

# Removes ticket data in bulk for selected booking and draw rows from the upload screen
@staff_member_required
@require_POST
def ticket_upload_bulk_delete(request):
    ids = request.POST.getlist("selected_ids")
    if not ids:
        messages.error(request, "No rows selected.")
        return redirect("ticket_upload")

    booking_ids = []
    draw_ids = []

    for v in ids:
        try:
            kind, pk = v.split(":", 1)
            if kind == "d":
                draw_ids.append(pk)
            else:
                booking_ids.append(pk)
        except ValueError:
            booking_ids.append(v)

    removed = 0
    missing = 0

    for obj in Booking.objects.filter(id__in=booking_ids):
        has_ticket = (
            obj.tickets.exists()
            or obj.ticket_type
            or obj.ticket_code
            or obj.ticket_sent
            or getattr(obj, "ticket_qr_value", "")
            or getattr(obj, "ticket_instructions", "")
            or getattr(obj, "generic_booking_code", "")
        )

        if not has_ticket:
            missing += 1
            continue

        obj.tickets.all().delete()
        obj.ticket_type = None
        obj.ticket_code = None
        obj.ticket_sent = False
        obj.ticket_sent_at = None
        obj.ticket_qr_value = ""
        obj.ticket_instructions = ""
        obj.generic_booking_code = ""
        obj.save()

        removed += 1

    for draw_obj in TicketDrawBooking.objects.filter(id__in=draw_ids).select_related("converted_booking"):
        obj = draw_obj.converted_booking

        if not obj:
            missing += 1
            continue

        has_ticket = (
            obj.tickets.exists()
            or obj.ticket_type
            or obj.ticket_code
            or obj.ticket_sent
            or getattr(obj, "ticket_qr_value", "")
            or getattr(obj, "ticket_instructions", "")
            or getattr(obj, "generic_booking_code", "")
        )

        if not has_ticket:
            missing += 1
            continue

        obj.tickets.all().delete()
        obj.ticket_type = None
        obj.ticket_code = None
        obj.ticket_sent = False
        obj.ticket_sent_at = None
        obj.ticket_qr_value = ""
        obj.ticket_instructions = ""
        obj.generic_booking_code = ""
        obj.save()

        removed += 1

    if removed:
        messages.success(request, f"Removed tickets for {removed} row(s).")
    if missing:
        messages.error(request, f"{missing} selected row(s) had no ticket to delete.")

    return redirect("ticket_upload")


@staff_member_required
def manage_terms_and_conditions(request):
    """
    Allow admins to edit Terms and Conditions sections through a web form
    with live preview.
    """
    from .models import TermsAndConditions
    
    t_and_c = TermsAndConditions.get()
    
    if request.method == 'POST':
        t_and_c.eligibility = request.POST.get('eligibility', '')
        t_and_c.application_limits = request.POST.get('application_limits', '')
        t_and_c.how_to_book = request.POST.get('how_to_book', '')
        t_and_c.attendance = request.POST.get('attendance', '')
        t_and_c.conduct = request.POST.get('conduct', '')
        t_and_c.liability_and_entry = request.POST.get('liability_and_entry', '')
        t_and_c.save()
        
        messages.success(request, "Terms and Conditions updated successfully.")
        return redirect('manage_terms_and_conditions')
    
    return render(request, 'fergusonbequest/manage_terms_and_conditions.html', {
        't_and_c': t_and_c,
    })
User = get_user_model()